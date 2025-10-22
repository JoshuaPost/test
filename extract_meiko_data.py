#!/usr/bin/env python3
"""
Extract Meiko data from HTML and populate Excel template
This creates a working example for the TP Dashboard Template
"""

import openpyxl

def extract_meiko_data_to_excel(template_file, output_file):
    """Extract data from Meiko.html and populate Excel template"""

    # Load the Excel template
    wb = openpyxl.load_workbook(template_file)

    # === Update Client Info ===
    ws_client = wb["Client Info"]
    ws_client['B2'] = 'Meiko Group'
    ws_client['B3'] = '2025-12-31'
    ws_client['B4'] = 'FYE 2025'

    # === Extract Countries Data ===
    ws_countries = wb["Countries"]

    # Define the countries data manually (extracted from your HTML)
    countries_data = [
        # Country Name, Country ID, Region, Entity Name, MF, LF, Forms, Thresholds, Documentation, Forms, Filing Status, Deadlines
        ['Germany', 'germany', 'Europe', 'Meiko Maschinenbau GmbH & Co. KG and local sales entities', 'Y', 'Y', 'N',
         'Master File required if German entity sales exceed EUR 100 million in the preceding FY. | Local File required if goods RPTs > EUR 6 million OR other RPTs > EUR 600,000.',
         'Maintain BEPS Action 13 compliant MF/LF with refreshed benchmarking. | Document extraordinary transactions within six months of year end.',
         'Transaction Matrix (NEW): Required structured overview of RPTs. | Master File, Transaction Matrix, and Extraordinary RPT Documentation must be submitted automatically upon audit announcement.',
         'Filing status: Automatic submission within 30 days of tax audit order; LF submitted upon request.',
         'Submission (Automatic): Within 30 days of receiving the tax audit order. | LF Submission (Upon Request): Within 30 days.'],

        ['Austria', 'austria', 'Europe', 'Meiko Austria GmbH', 'Y', 'Y', 'N',
         'MF/LF mandatory if local GAAP turnover exceeded EUR 50M in each of the two preceding financial years.',
         'Prepare MF/LF aligned with OECD BEPS Action 13 standards.',
         'CbCR notification via FinanzOnline portal.',
         'Filing status: MF/LF submission only upon request.',
         'CbCR notification: 31 Dec 2025 | MF/LF submission: Upon request, within 30 days after filing the tax return.'],

        ['Belgium', 'belgium', 'Europe', 'Meiko Belgium', 'Y', 'Y', 'Y',
         'MF required if consolidated group revenue exceeds EUR 750M. | LF required if annual Belgian turnover > EUR 50M OR annual RPTs > EUR 1M.',
         'Maintain OECD-compliant MF/LF documentation.',
         'Form 275.CBC.NOT (CbCR notification) | Form 275.LF (Local File summary) | Form 275.MF (Master File summary)',
         'Filing status: Forms filed; full documentation upon request.',
         'Form 275.CBC.NOT: 31 Dec 2025 | Form 275.LF with CIT return (expected 31 Jul 2026) | Form 275.MF: 31 Dec 2026'],

        ['France', 'france', 'Europe', 'Meiko France', 'Y', 'Y', 'Y',
         'Master File required if consolidated revenue > EUR 750M. | Local File required if entity turnover > EUR 50M OR RPTs > EUR 100M.',
         'Prepare comprehensive OECD-compliant Master File and Local File.',
         'CERFA 2257-SD form (TP Schedules)',
         'Filing status: Forms filed annually.',
         'File CERFA 2257-SD by 31 Oct 2026 (extended deadline for large businesses).'],

        ['Netherlands', 'netherlands', 'Europe', 'Meiko Netherlands', 'Y', 'Y', 'N',
         'MF required if ultimate parent has consolidated revenue > EUR 50M. | LF required if local entity meets threshold.',
         'Maintain OECD BEPS Action 13 compliant documentation.',
         'Public CbCR report (if revenue > EUR 750M)',
         'Filing status: MF/LF submission upon request; Public CbCR published.',
         'MF/LF preparation: By CIT return deadline (Standard: 31 May 2026) | MF/LF submission: Upon request, due within 30 days.'],

        ['Spain', 'spain', 'Europe', 'Meiko Spain', 'Y', 'Y', 'Y',
         'Master File if group consolidated turnover > EUR 45M. | Local File required if net turnover > EUR 45M OR RPT transactions > EUR 250,000.',
         'Maintain comprehensive TP documentation.',
         'Form 232 (TP informative return)',
         'Filing status: Form 232 filed annually; MF/LF upon request (10-day audit window).',
         'MF/LF ready by the CIT filing date (approx. 25 Jul 2026). | File Form 232 by approx. 25 Aug 2026.'],

        ['Turkey', 'turkey', 'Europe', 'Meiko Turkey', 'Y', 'Y', 'Y',
         'MF/LF required if group consolidated revenue > TRY 500M (≈ EUR 16M) AND local entity revenue > TRY 200M (≈ EUR 6.5M).',
         'Maintain comprehensive TP documentation.',
         'Form F 982 (TP informative return)',
         'Filing status: TP return filed with CIT; documentation upon request.',
         'File TP return (Form F 982) with the CIT return by 31 Mar 2026. | MF/LF preparation: 12 months after FYE (31 Dec 2026).'],

        ['United Kingdom', 'uk', 'Europe', 'Meiko UK Ltd.', 'Y', 'Y', 'Y',
         'MF/LF required if group revenue > £750M OR complex RPTs. | Recommended for all groups as penalty protection.',
         'Maintain TP documentation aligned with OECD standards.',
         'CIT return includes TP disclosure.',
         'Filing status: Submission only upon request.',
         'File the CIT return including TP disclosure by 30 Sep 2026.'],

        ['Italy', 'italy', 'Europe', 'Meiko Italy', 'Y', 'Y', 'Y',
         'MF required if parent prepares MF for group. | LF required if certain thresholds are met.',
         'Maintain comprehensive TP documentation aligned with OECD standards.',
         'RS 106 disclosure form',
         'Filing status: Documentation upon request; must be properly timestamped.',
         'Complete and timestamp MF/LF documentation by 31 Oct 2026 (before RS 106 filing).'],

        ['Poland', 'poland', 'Europe', 'Meiko Poland', 'Y', 'Y', 'Y',
         'MF/LF required if group consolidated revenue > EUR 750M OR local entity revenue > EUR 2M.',
         'Prepare comprehensive TP documentation.',
         'TPR-C form (TP informative return)',
         'Filing status: TPR-C form filed; documentation upon request.',
         'Complete Local File preparation by 31 Oct 2026. | File TPR-C form by 30 Nov 2026. | Complete Master File preparation by 31 Dec 2026.'],

        ['Serbia', 'serbia', 'Europe', 'Meiko Serbia', 'Y', 'Y', 'Y',
         'MF/LF required for all entities with RPTs.',
         'Maintain OECD-compliant TP documentation.',
         'TP Documentation submitted with CIT return',
         'Filing status: TP Documentation submitted with CIT return.',
         'Submit TP Documentation with CIT return by 29 Jun 2026.'],

        ['Switzerland', 'switzerland', 'Europe', 'Meiko Switzerland', 'Y', 'Y', 'N',
         'TP documentation recommended for all entities with RPTs.',
         'Maintain TP documentation for penalty protection.',
         'No specific forms required.',
         'Filing status: Documentation provided upon request during audit.',
         'Provide documentation within 30 days upon request.'],

        ['Mexico', 'mexico', 'Americas', 'Meiko Mexico', 'Y', 'Y', 'Y',
         'Master File required if group revenue > MXN 8.1 billion (≈ EUR 385M). | Local File required for all entities.',
         'Prepare comprehensive TP documentation.',
         'Master File informative return | Local File informative return',
         'Filing status: Informative returns filed annually.',
         'Prepare the Local File by 31 Mar 2026 (or 15 May 2026 if ISSIF is filed). | File the Local File informative return by 15 May 2026. | File the Master File informative return by 31 Dec 2026.'],

        ['Canada', 'canada', 'Americas', 'Meiko Canada', 'Y', 'Y', 'Y',
         'TP documentation required for all entities with RPTs > CAD 1M.',
         'Maintain comprehensive TP documentation.',
         'Form T-106 (Information Return of Non-Arm\'s Length Transactions)',
         'Filing status: Form T-106 filed annually.',
         'Complete TP documentation and file T-106 by 30 Jun 2026.'],

        ['United States', 'us', 'Americas', 'Meiko US', 'Y', 'Y', 'N',
         'TP documentation recommended for penalty protection.',
         'Maintain contemporaneous TP documentation.',
         'No specific TP forms required.',
         'Filing status: Documentation upon request.',
         'Submission: Upon request, typically within 30 days.'],

        ['Japan', 'japan', 'Asia-Pacific', 'Meiko Japan KK', 'Y', 'Y', 'Y',
         'Master File required if group revenue > JPY 100 billion (≈ EUR 611M). | Local File required if RPTs > JPY 5 billion (≈ EUR 31M).',
         'Prepare OECD-compliant TP documentation.',
         'Schedule 17-4 (RPT disclosure)',
         'Filing status: Schedule 17-4 filed with tax return; MF/LF upon request.',
         'Submit Schedule 17-4 with the return and maintain the Local File by 31 Mar 2026.'],

        ['India', 'india', 'Asia-Pacific', 'Meiko Clean Solutions (India) Pvt. Ltd. (FYE: Mar 31, 2026)', 'Y', 'Y', 'Y',
         'Master File required if group revenue > INR 50 billion (≈ EUR 540M). | Local File required if RPTs > INR 10 crore (≈ EUR 1.1M).',
         'Maintain comprehensive TP documentation.',
         'Form 3CEB (TP audit report)',
         'Filing status: Form 3CEB filed annually.',
         'Documentation completed by 31 Oct 2026. | File Form 3CEB by 30 Nov 2026.'],

        ['Australia', 'australia', 'Asia-Pacific', 'Meiko Australia', 'Y', 'Y', 'Y',
         'TP documentation required for entities with RPTs > AUD 2M (≈ EUR 1.2M).',
         'Maintain OECD-compliant TP documentation.',
         'International Dealings Schedule (IDS) filed with tax return',
         'Filing status: IDS filed annually; documentation upon request.',
         'Documentation ready by the CIT lodgment date (15 Jul 2026). | Lodge CIT return with IDS by 15 Jul 2026.'],

        ['China', 'china', 'Asia-Pacific', 'Meiko China', 'Y', 'Y', 'Y',
         'Master File required if annual cross-border RPTs exceed RMB 1 billion (≈ EUR 130 million) OR the UHC prepared an MF. | Local File required if annual RPTs exceed certain thresholds.',
         'Maintain detailed Local File study supporting arm\'s-length pricing. | Ensure comprehensive documentation for all related party transactions.',
         'Annual RPT Reporting Forms: Multiple detailed forms/questionnaires filed annually.',
         'Filing status: RPT forms filed with annual tax filing.',
         'RPT Forms submission: Filed with the annual tax filing on or before 31 May 2026. | Local File preparation: Must be ready by 30 Jun 2026.'],

        ['Hong Kong', 'hongkong', 'Asia-Pacific', 'Meiko Hong Kong', 'Y', 'Y', 'N',
         'MF/LF required based on local entity size/RPT volume criteria.',
         'Maintain TP documentation aligned with OECD standards. | Prepare comprehensive analysis supporting transfer pricing positions.',
         'Supplementary Form S2 filed with the Profits Tax Return. | Form IR 1475 (requested from selected taxpayers).',
         'Filing status: Submission only upon request.',
         'Preparation: Within 9 months after FYE (30 Sep 2026). | Submission: Only upon request.'],

        ['Malaysia', 'malaysia', 'Asia-Pacific', 'Meiko Malaysia', 'N', 'Y', 'N',
         'Master File not mandatory (Group revenue below MYR 3 billion threshold ≈ EUR 608 million). MF content integrated into Local File (CTPD). | Local File required if Gross Income > MYR 30 million (≈ EUR 6.08 million) AND cross-border RPTs total > MYR 10 million (≈ EUR 2.03 million).',
         'Prepare comprehensive Contemporaneous Transfer Pricing Documentation (CTPD). | Integrate Master File content into Local File.',
         'Disclosure in Form C: Taxpayers must declare (via check box) in the CIT return whether they have prepared TP documentation.',
         'Filing status: Declaration in CIT return; documentation submitted upon request.',
         'Preparation: Prior to the due date for furnishing the tax return (7 months after FYE: 31 Jul 2026). | Submission: Upon request, within 14 days.'],

        ['United Arab Emirates', 'uae', 'Middle East', 'Meiko UAE', 'Y', 'Y', 'N',
         'TP documentation required for all entities subject to CT with RPTs.',
         'Maintain comprehensive TP documentation.',
         'No specific TP forms required.',
         'Filing status: Documentation upon request.',
         'Documentation recommended by the CIT deadline (30 Sep 2026). | Submission: Upon request.']
    ]

    # Write countries data to Excel
    row_num = 3  # Start after header and instruction rows
    for country in countries_data:
        for col_num, value in enumerate(country, start=1):
            ws_countries.cell(row=row_num, column=col_num, value=value)
        row_num += 1

    # === Extract Timeline Data ===
    ws_timeline = wb["Timeline"]

    timeline_data = [
        # Quarter, Country, Date, Description, Type
        ['December 2025', 'Austria', '2025-12-31', 'CbCR notification via FinanzOnline portal', 'filing'],
        ['December 2025', 'Belgium', '2025-12-31', 'File Form 275.CBC.NOT (CbCR notification)', 'filing'],

        ['Q1 2026 (January - March)', 'Austria', None, 'MF/LF submission within 30 days after filing tax return', 'upon-request'],
        ['Q1 2026 (January - March)', 'Japan', '2026-03-31', 'Submit Schedule 17-4 with the return and maintain the Local File', 'filing'],
        ['Q1 2026 (January - March)', 'Mexico', '2026-03-31', 'Prepare the Local File (or 15 May 2026 if ISSIF is filed)', 'preparation'],
        ['Q1 2026 (January - March)', 'Turkey', '2026-03-31', 'File TP return (Form F 982) with the CIT return', 'filing'],

        ['Q2 2026 (April - June)', 'Mexico', '2026-05-15', 'File the Local File informative return', 'filing'],
        ['Q2 2026 (April - June)', 'China', '2026-05-31', 'File RPT Forms with annual tax filing', 'filing'],
        ['Q2 2026 (April - June)', 'Netherlands', '2026-05-31', 'Complete MF/LF preparation for penalty protection', 'preparation'],
        ['Q2 2026 (April - June)', 'Canada', '2026-06-30', 'Complete TP documentation and file T-106', 'filing'],
        ['Q2 2026 (April - June)', 'China', '2026-06-30', 'Complete Local File preparation for penalty protection', 'preparation'],
        ['Q2 2026 (April - June)', 'Germany', '2026-06-30', 'Extraordinary transaction documentation due', 'filing'],
        ['Q2 2026 (April - June)', 'Serbia', '2026-06-29', 'Submit TP Documentation with CIT return', 'filing'],

        ['Q3 2026 (July - September)', 'Australia', '2026-07-15', 'Complete TP documentation and lodge CIT return with IDS', 'filing'],
        ['Q3 2026 (July - September)', 'Spain', '2026-07-25', 'Complete MF/LF preparation for penalty protection', 'preparation'],
        ['Q3 2026 (July - September)', 'Belgium', '2026-07-31', 'File Form 275.LF with the CIT return (expected)', 'filing'],
        ['Q3 2026 (July - September)', 'Malaysia', '2026-07-31', 'Complete CTPD documentation for penalty protection', 'preparation'],
        ['Q3 2026 (July - September)', 'Spain', '2026-08-25', 'File Form 232 (approx.)', 'filing'],
        ['Q3 2026 (July - September)', 'United Kingdom', '2026-09-30', 'File the CIT return including TP disclosure', 'filing'],
        ['Q3 2026 (July - September)', 'UAE', '2026-09-30', 'Complete TP documentation for penalty protection', 'preparation'],
        ['Q3 2026 (July - September)', 'Hong Kong', '2026-09-30', 'Complete TP documentation for penalty protection', 'preparation'],

        ['Q4 2026 (October - December)', 'France', '2026-10-31', 'File CERFA 2257-SD', 'filing'],
        ['Q4 2026 (October - December)', 'India', '2026-10-31', 'Complete TP documentation for penalty protection', 'preparation'],
        ['Q4 2026 (October - December)', 'Italy', '2026-10-31', 'Complete and timestamp MF/LF documentation', 'preparation'],
        ['Q4 2026 (October - December)', 'Poland', '2026-10-31', 'Complete Local File preparation', 'preparation'],
        ['Q4 2026 (October - December)', 'India', '2026-11-30', 'File Form 3CEB', 'filing'],
        ['Q4 2026 (October - December)', 'Poland', '2026-11-30', 'File TPR-C form', 'filing'],
        ['Q4 2026 (October - December)', 'Mexico', '2026-12-31', 'File the Master File informative return', 'filing'],
        ['Q4 2026 (October - December)', 'Belgium', '2026-12-31', 'File Form 275.MF', 'filing'],
        ['Q4 2026 (October - December)', 'Poland', '2026-12-31', 'Complete Master File preparation', 'preparation'],
        ['Q4 2026 (October - December)', 'Turkey', '2026-12-31', 'Complete MF/LF preparation', 'preparation']
    ]

    # Write timeline data to Excel
    row_num = 3  # Start after header and instruction rows
    for item in timeline_data:
        for col_num, value in enumerate(item, start=1):
            ws_timeline.cell(row=row_num, column=col_num, value=value)
        row_num += 1

    # Save the populated workbook
    wb.save(output_file)
    print(f"✓ Excel file populated with Meiko data: {output_file}")
    print(f"  - {len(countries_data)} countries added")
    print(f"  - {len(timeline_data)} timeline items added")


if __name__ == "__main__":
    extract_meiko_data_to_excel(
        '/home/user/test/TP_Dashboard_Template.xlsx',
        '/home/user/test/Meiko_TP_Data.xlsx'
    )
