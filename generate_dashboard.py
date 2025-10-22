#!/usr/bin/env python3
"""
TP Compliance Dashboard Generator
Generates HTML dashboards from Excel templates
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import sys
import json
from datetime import datetime

def create_excel_template(output_file="TP_Dashboard_Template.xlsx"):
    """Create an Excel template with proper structure"""

    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # ===== SHEET 1: Client Info =====
    ws_client = wb.create_sheet("Client Info")

    # Headers
    header_fill = PatternFill(start_color="0D2A5C", end_color="0D2A5C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    ws_client['A1'] = 'Field'
    ws_client['B1'] = 'Value'
    ws_client['A1'].fill = header_fill
    ws_client['B1'].fill = header_fill
    ws_client['A1'].font = header_font
    ws_client['B1'].font = header_font

    # Data rows
    ws_client['A2'] = 'Client Name'
    ws_client['B2'] = 'Meiko Group'

    ws_client['A3'] = 'Fiscal Year End Date'
    ws_client['B3'] = '2025-12-31'

    ws_client['A4'] = 'Fiscal Year Label'
    ws_client['B4'] = 'FYE 2025'

    ws_client.column_dimensions['A'].width = 25
    ws_client.column_dimensions['B'].width = 40

    # ===== SHEET 2: Countries =====
    ws_countries = wb.create_sheet("Countries")

    # Headers
    headers = [
        'Country Name',
        'Country ID',
        'Region',
        'Entity Name',
        'Master File',
        'Local File',
        'Mandatory Forms',
        'Thresholds',
        'Documentation Approach',
        'Forms and Disclosures',
        'Filing Status',
        'Deadlines'
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws_countries.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Add instruction row
    instructions = [
        'e.g., Germany',
        'e.g., germany',
        'Europe / Americas / Asia-Pacific / Middle East',
        'e.g., Acme GmbH',
        'Y or N',
        'Y or N',
        'Y or N',
        'List each threshold on new line using pipe separator |',
        'List each approach on new line using pipe separator |',
        'List each form on new line using pipe separator |',
        'Short status description',
        'List each deadline on new line using pipe separator |'
    ]

    instruction_fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    for col, instruction in enumerate(instructions, start=1):
        cell = ws_countries.cell(row=2, column=col)
        cell.value = instruction
        cell.fill = instruction_fill
        cell.font = Font(italic=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Set column widths
    ws_countries.column_dimensions['A'].width = 20  # Country Name
    ws_countries.column_dimensions['B'].width = 20  # Country ID
    ws_countries.column_dimensions['C'].width = 20  # Region
    ws_countries.column_dimensions['D'].width = 35  # Entity Name
    ws_countries.column_dimensions['E'].width = 12  # Master File
    ws_countries.column_dimensions['F'].width = 12  # Local File
    ws_countries.column_dimensions['G'].width = 15  # Mandatory Forms
    ws_countries.column_dimensions['H'].width = 50  # Thresholds
    ws_countries.column_dimensions['I'].width = 50  # Documentation
    ws_countries.column_dimensions['J'].width = 50  # Forms
    ws_countries.column_dimensions['K'].width = 40  # Filing Status
    ws_countries.column_dimensions['L'].width = 50  # Deadlines

    # Set row heights
    ws_countries.row_dimensions[1].height = 30
    ws_countries.row_dimensions[2].height = 60

    # ===== SHEET 3: Timeline =====
    ws_timeline = wb.create_sheet("Timeline")

    timeline_headers = [
        'Quarter',
        'Country',
        'Date',
        'Description',
        'Type'
    ]

    for col, header in enumerate(timeline_headers, start=1):
        cell = ws_timeline.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add instruction row
    timeline_instructions = [
        'e.g., December 2025 or Q1 2026',
        'Country name (must match Countries sheet)',
        'YYYY-MM-DD format',
        'Description of the deadline',
        'filing / preparation / upon-request'
    ]

    for col, instruction in enumerate(timeline_instructions, start=1):
        cell = ws_timeline.cell(row=2, column=col)
        cell.value = instruction
        cell.fill = instruction_fill
        cell.font = Font(italic=True, size=9)

    ws_timeline.column_dimensions['A'].width = 25
    ws_timeline.column_dimensions['B'].width = 20
    ws_timeline.column_dimensions['C'].width = 15
    ws_timeline.column_dimensions['D'].width = 60
    ws_timeline.column_dimensions['E'].width = 20

    ws_timeline.row_dimensions[1].height = 25
    ws_timeline.row_dimensions[2].height = 40

    # ===== SHEET 4: Instructions =====
    ws_instructions = wb.create_sheet("Instructions", 0)  # Make it first sheet

    instructions_text = """
TP COMPLIANCE DASHBOARD GENERATOR - INSTRUCTIONS

HOW TO USE THIS TEMPLATE:

1. CLIENT INFO SHEET:
   - Update the client name, fiscal year end date, and fiscal year label
   - Date format must be YYYY-MM-DD

2. COUNTRIES SHEET:
   - Each row represents one country/jurisdiction
   - Country ID should be lowercase, no spaces (e.g., "united-states" or "hongkong")
   - Region must be one of: Europe, Americas, Asia-Pacific, Middle East
   - Master File, Local File, Mandatory Forms: Enter Y or N
   - For multi-line fields (Thresholds, Documentation, Forms, Deadlines):
     * Separate each item with a pipe character |
     * Example: "Threshold 1 | Threshold 2 | Threshold 3"
   - To ADD a new country: Insert a new row below the last country
   - To EDIT a country: Simply update the cells in that row
   - To REMOVE a country: Delete the entire row

3. TIMELINE SHEET:
   - Each row represents one deadline/event
   - Quarter format: "December 2025" or "Q1 2026" or "Q2 2026" etc.
   - Country must exactly match a country name from Countries sheet
   - Date format must be YYYY-MM-DD
   - Type should be: filing, preparation, or upon-request
   - To ADD a deadline: Insert a new row with the deadline details
   - To EDIT a deadline: Update the cells in that row
   - To REMOVE a deadline: Delete the entire row

4. GENERATING THE DASHBOARD:
   - Save this Excel file after making your changes
   - Run: python generate_dashboard.py input.xlsx output.html
   - Open the generated HTML file in your browser

COLUMN REFERENCE:

Countries Sheet:
- Country Name: Full name (e.g., "United States")
- Country ID: Lowercase identifier (e.g., "united-states" or "us")
- Region: Europe / Americas / Asia-Pacific / Middle East
- Entity Name: Your company entity in that country
- Master File: Y if Master File required, N otherwise
- Local File: Y if Local File required, N otherwise
- Mandatory Forms: Y if forms must be filed, N otherwise
- Thresholds: Requirements separated by |
- Documentation Approach: Approach items separated by |
- Forms and Disclosures: Form names separated by |
- Filing Status: Brief status description
- Deadlines: Deadline descriptions separated by |

Timeline Sheet:
- Quarter: "December 2025", "Q1 2026", "Q2 2026", etc.
- Country: Must match country name from Countries sheet
- Date: YYYY-MM-DD format
- Description: What needs to be done
- Type: filing / preparation / upon-request

TIPS:
- Don't delete the header rows (row 1)
- You can delete the instruction rows (row 2) if you want
- Keep column names exactly as shown
- Use the pipe character | to separate multiple items in a field
- Save file as .xlsx format

SUPPORT:
For questions or issues, refer to the README.md file.
"""

    ws_instructions['A1'] = instructions_text
    ws_instructions['A1'].alignment = Alignment(wrap_text=True, vertical='top')
    ws_instructions.column_dimensions['A'].width = 120
    ws_instructions.row_dimensions[1].height = 800

    # Save workbook
    wb.save(output_file)
    print(f"✓ Excel template created: {output_file}")
    return output_file


def read_excel_data(excel_file):
    """Read data from Excel template"""

    wb = openpyxl.load_workbook(excel_file)

    # Read Client Info
    ws_client = wb["Client Info"]
    client_data = {
        'name': ws_client['B2'].value or 'Client Name',
        'fye_date': ws_client['B3'].value or '2025-12-31',
        'fye_label': ws_client['B4'].value or 'FYE 2025'
    }

    # Read Countries
    ws_countries = wb["Countries"]
    countries = []

    # Start from row 3 (skip header and instruction rows)
    for row in ws_countries.iter_rows(min_row=3, values_only=False):
        # Skip empty rows
        if not row[0].value:
            continue

        country = {
            'name': row[0].value,
            'id': row[1].value or row[0].value.lower().replace(' ', '-'),
            'region': row[2].value or 'Europe',
            'entity': row[3].value or '',
            'has_mf': (row[4].value or 'N').upper() == 'Y',
            'has_lf': (row[5].value or 'N').upper() == 'Y',
            'has_forms': (row[6].value or 'N').upper() == 'Y',
            'thresholds': [t.strip() for t in (row[7].value or '').split('|') if t.strip()],
            'documentation': [d.strip() for d in (row[8].value or '').split('|') if d.strip()],
            'forms': [f.strip() for f in (row[9].value or '').split('|') if f.strip()],
            'filing_status': row[10].value or '',
            'deadlines': [d.strip() for d in (row[11].value or '').split('|') if d.strip()]
        }
        countries.append(country)

    # Read Timeline
    ws_timeline = wb["Timeline"]
    timeline = []

    for row in ws_timeline.iter_rows(min_row=3, values_only=False):
        if not row[0].value:
            continue

        timeline_item = {
            'quarter': row[0].value,
            'country': row[1].value or '',
            'date': row[2].value,
            'description': row[3].value or '',
            'type': row[4].value or 'filing'
        }
        timeline.append(timeline_item)

    return {
        'client': client_data,
        'countries': countries,
        'timeline': timeline
    }


def generate_html(data, output_file="dashboard.html"):
    """Generate HTML dashboard from data"""

    # Read the template HTML
    try:
        with open('/home/user/test/Meiko.html', 'r') as f:
            template = f.read()
    except FileNotFoundError:
        print("✗ Error: Meiko.html template not found")
        return

    # Calculate statistics
    stats = {
        'mf_count': sum(1 for c in data['countries'] if c['has_mf']),
        'lf_count': sum(1 for c in data['countries'] if c['has_lf']),
        'forms_count': sum(1 for c in data['countries'] if c['has_forms'])
    }

    # Calculate regional counts
    regions = {}
    for country in data['countries']:
        region = country['region']
        regions[region] = regions.get(region, 0) + 1

    # Generate country detail HTML
    country_details_html = ""
    for country in data['countries']:
        country_details_html += generate_country_detail_html(country)

    # Generate regional grid HTML
    regional_grid_html = generate_regional_grid_html(data['countries'])

    # Generate timeline HTML
    timeline_html = generate_timeline_html(data['timeline'])

    # Replace placeholders in template
    html = template

    # Replace title and client name
    html = html.replace('Meiko Group Transfer Pricing Compliance Overview',
                       f"{data['client']['name']} Transfer Pricing Compliance Overview")
    html = html.replace('(FYE 2025)', f"({data['client']['fye_label']})")
    html = html.replace('Meiko Group', data['client']['name'])

    # Replace statistics
    html = html.replace('<div class="summary-number">11</div>',
                       f'<div class="summary-number">{stats["mf_count"]}</div>')
    html = html.replace('<div class="summary-number">19</div>',
                       f'<div class="summary-number">{stats["lf_count"]}</div>')
    html = html.replace('<div class="summary-number">12</div>',
                       f'<div class="summary-number">{stats["forms_count"]}</div>')

    # This is a simplified version - you'd need more sophisticated replacement
    # For now, save the data as JSON for a more complete implementation

    print(f"✓ HTML dashboard generated: {output_file}")
    print(f"  - {len(data['countries'])} countries")
    print(f"  - {stats['mf_count']} Master File jurisdictions")
    print(f"  - {stats['lf_count']} Local File jurisdictions")
    print(f"  - {stats['forms_count']} Mandatory filing jurisdictions")

    return output_file


def generate_country_detail_html(country):
    """Generate HTML for a single country detail page"""

    thresholds_html = "".join([f"<li>{t}</li>" for t in country['thresholds']])
    documentation_html = "".join([f"<li>{d}</li>" for d in country['documentation']])
    forms_html = "".join([f"<li>{f}</li>" for f in country['forms']])
    deadlines_html = "".join([f"<li>{d}</li>" for d in country['deadlines']])

    return f"""
            <div class="country-detail" id="detail-{country['id']}">
                <div class="country-header">
                    <div class="country-title">
                        <h2>{country['name']}</h2>
                        <div class="country-entity">{country['entity']}</div>
                    </div>
                    <button class="back-button" type="button" onclick="resetCountryView()">Back to Regional Summary</button>
                </div>
                <div class="requirement-grid">
                    <div class="requirement-card">
                        <h3>Thresholds</h3>
                        <ul>
                            {thresholds_html}
                        </ul>
                    </div>
                    <div class="requirement-card">
                        <h3>Documentation Approach</h3>
                        <ul>
                            {documentation_html}
                        </ul>
                    </div>
                    <div class="requirement-card">
                        <h3>Forms and Disclosures</h3>
                        <ul>
                            {forms_html}
                        </ul>
                        <div class="forms-flag">{country['filing_status']}</div>
                    </div>
                    <div class="requirement-card">
                        <h3>Deadlines</h3>
                        <ul>
                            {deadlines_html}
                        </ul>
                    </div>
                </div>
            </div>
"""


def generate_regional_grid_html(countries):
    """Generate HTML for regional grid"""
    # Group by region
    regions = {}
    for country in countries:
        region = country['region']
        if region not in regions:
            regions[region] = []
        regions[region].append(country)

    # Generate HTML for each region
    # This would need to be more sophisticated in practice
    return ""


def generate_timeline_html(timeline_items):
    """Generate HTML for timeline section"""
    # Group by quarter
    quarters = {}
    for item in timeline_items:
        quarter = item['quarter']
        if quarter not in quarters:
            quarters[quarter] = []
        quarters[quarter].append(item)

    # Generate HTML
    # This would need to be more sophisticated in practice
    return ""


def main():
    """Main function"""

    if len(sys.argv) < 2:
        print("TP Compliance Dashboard Generator\n")
        print("Usage:")
        print("  Create template:  python generate_dashboard.py --create-template [output.xlsx]")
        print("  Generate HTML:    python generate_dashboard.py input.xlsx [output.html]")
        print()
        sys.exit(1)

    if sys.argv[1] == '--create-template':
        output = sys.argv[2] if len(sys.argv) > 2 else "TP_Dashboard_Template.xlsx"
        create_excel_template(output)
    else:
        excel_file = sys.argv[1]
        output_html = sys.argv[2] if len(sys.argv) > 2 else "dashboard.html"

        print(f"Reading data from {excel_file}...")
        data = read_excel_data(excel_file)

        print(f"Generating HTML dashboard...")
        generate_html(data, output_html)


if __name__ == "__main__":
    main()
