#!/usr/bin/env python3
"""
Create Excel templates for rule-based TP compliance system
Handles multiple conditions with AND/OR logic
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_country_rules_library():
    """Create the Country Rules Library template"""

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Styling
    header_fill = PatternFill(start_color="0D2A5C", end_color="0D2A5C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    instruction_fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    example_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    # ===== SHEET 1: Master File Requirements =====
    ws_mf = wb.create_sheet("MF Requirements")

    headers = [
        'Rule ID',
        'Country',
        'Condition Group',
        'Group Logic',
        'Metric Type',
        'Metric Scope',
        'Threshold Value',
        'Currency',
        'Operator',
        'Notes'
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws_mf.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Instructions row
    instructions = [
        'Unique ID per rule set',
        'Country name',
        'Group number (conditions in same group use OR)',
        'OR / AND between groups',
        'Revenue / Employees / Balance Sheet / RPTs / Always',
        'Group / Local Entity / Transaction',
        'Numeric threshold',
        'EUR / USD / GBP / etc',
        '>= / > / = / < / <=',
        'Additional context'
    ]

    for col, instr in enumerate(instructions, start=1):
        cell = ws_mf.cell(row=2, column=col)
        cell.value = instr
        cell.fill = instruction_fill
        cell.font = Font(italic=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Example: Germany MF (Group revenue >= 750M OR Local sales >= 100M)
    examples = [
        ['MF-DE-1', 'Germany', '1', 'OR', 'Revenue', 'Group (Consolidated)', 750000000, 'EUR', '>=', 'Standard BEPS threshold'],
        ['MF-DE-1', 'Germany', '2', 'OR', 'Revenue', 'Local Entity', 100000000, 'EUR', '>=', 'German entity sales threshold'],
        ['MF-ES-1', 'Spain', '1', 'OR', 'Revenue', 'Group (Consolidated)', 45000000, 'EUR', '>', 'Group consolidated turnover'],
    ]

    for row_idx, example in enumerate(examples, start=3):
        for col_idx, value in enumerate(example, start=1):
            cell = ws_mf.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.fill = example_fill

    # Set column widths
    ws_mf.column_dimensions['A'].width = 12
    ws_mf.column_dimensions['B'].width = 15
    ws_mf.column_dimensions['C'].width = 15
    ws_mf.column_dimensions['D'].width = 12
    ws_mf.column_dimensions['E'].width = 20
    ws_mf.column_dimensions['F'].width = 25
    ws_mf.column_dimensions['G'].width = 18
    ws_mf.column_dimensions['H'].width = 10
    ws_mf.column_dimensions['I'].width = 10
    ws_mf.column_dimensions['J'].width = 40

    ws_mf.row_dimensions[1].height = 30
    ws_mf.row_dimensions[2].height = 60

    # ===== SHEET 2: Local File Requirements =====
    ws_lf = wb.create_sheet("LF Requirements")

    for col, header in enumerate(headers, start=1):
        cell = ws_lf.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col, instr in enumerate(instructions, start=1):
        cell = ws_lf.cell(row=2, column=col)
        cell.value = instr
        cell.fill = instruction_fill
        cell.font = Font(italic=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Example: Germany LF (Goods RPTs > 6M OR Other RPTs > 600K)
    examples_lf = [
        ['LF-DE-1', 'Germany', '1', 'OR', 'RPTs', 'Transaction (Goods)', 6000000, 'EUR', '>', 'Goods-related transactions'],
        ['LF-DE-1', 'Germany', '2', 'OR', 'RPTs', 'Transaction (Other)', 600000, 'EUR', '>', 'Other related party transactions'],
        ['LF-ES-1', 'Spain', '1', 'OR', 'Revenue', 'Local Entity', 45000000, 'EUR', '>', 'Net turnover threshold'],
        ['LF-ES-1', 'Spain', '2', 'OR', 'RPTs', 'Transaction (All)', 250000, 'EUR', '>', 'Related party transactions'],
    ]

    for row_idx, example in enumerate(examples_lf, start=3):
        for col_idx, value in enumerate(example, start=1):
            cell = ws_lf.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.fill = example_fill

    # Copy column widths
    for col in range(1, 11):
        ws_lf.column_dimensions[get_column_letter(col)].width = ws_mf.column_dimensions[get_column_letter(col)].width

    ws_lf.row_dimensions[1].height = 30
    ws_lf.row_dimensions[2].height = 60

    # ===== SHEET 3: CbCR Requirements =====
    ws_cbcr = wb.create_sheet("CbCR Requirements")

    for col, header in enumerate(headers, start=1):
        cell = ws_cbcr.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col, instr in enumerate(instructions, start=1):
        cell = ws_cbcr.cell(row=2, column=col)
        cell.value = instr
        cell.fill = instruction_fill
        cell.font = Font(italic=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    examples_cbcr = [
        ['CBCR-DE-1', 'Germany', '1', 'OR', 'Revenue', 'Group (Consolidated)', 750000000, 'EUR', '>=', 'Standard OECD threshold'],
        ['CBCR-ES-1', 'Spain', '1', 'OR', 'Revenue', 'Group (Consolidated)', 750000000, 'EUR', '>=', 'Standard OECD threshold'],
    ]

    for row_idx, example in enumerate(examples_cbcr, start=3):
        for col_idx, value in enumerate(example, start=1):
            cell = ws_cbcr.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.fill = example_fill

    for col in range(1, 11):
        ws_cbcr.column_dimensions[get_column_letter(col)].width = ws_mf.column_dimensions[get_column_letter(col)].width

    ws_cbcr.row_dimensions[1].height = 30
    ws_cbcr.row_dimensions[2].height = 60

    # ===== SHEET 4: Forms Requirements =====
    ws_forms = wb.create_sheet("Forms Requirements")

    forms_headers = [
        'Country',
        'Form Name',
        'Form Type',
        'Condition Logic',
        'Filing Deadline Rule',
        'Notes'
    ]

    for col, header in enumerate(forms_headers, start=1):
        cell = ws_forms.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    forms_instructions = [
        'Country name',
        'Official form name/number',
        'TP Return / Disclosure / Notification / CbCR',
        'Always / If MF Required / If LF Required / Custom',
        'Deadline calculation rule (e.g., CIT+30days, FYE+10months)',
        'Additional context'
    ]

    for col, instr in enumerate(forms_instructions, start=1):
        cell = ws_forms.cell(row=2, column=col)
        cell.value = instr
        cell.fill = instruction_fill
        cell.font = Font(italic=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    forms_examples = [
        ['Germany', 'Transaction Matrix', 'TP Disclosure', 'Always', 'Upon audit (30 days)', 'NEW 2024 requirement'],
        ['Spain', 'Form 232', 'TP Return', 'If MF or LF Required', 'CIT filing + 30 days', 'Annual informative return'],
    ]

    for row_idx, example in enumerate(forms_examples, start=3):
        for col_idx, value in enumerate(example, start=1):
            cell = ws_forms.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.fill = example_fill

    ws_forms.column_dimensions['A'].width = 15
    ws_forms.column_dimensions['B'].width = 25
    ws_forms.column_dimensions['C'].width = 20
    ws_forms.column_dimensions['D'].width = 25
    ws_forms.column_dimensions['E'].width = 30
    ws_forms.column_dimensions['F'].width = 40

    ws_forms.row_dimensions[1].height = 30
    ws_forms.row_dimensions[2].height = 60

    # ===== SHEET 5: Deadlines =====
    ws_deadlines = wb.create_sheet("Deadlines")

    deadline_headers = [
        'Country',
        'Requirement Type',
        'Description',
        'Deadline Rule',
        'Fixed Date',
        'Offset Type',
        'Offset Value',
        'Notes'
    ]

    for col, header in enumerate(deadline_headers, start=1):
        cell = ws_deadlines.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    deadline_instructions = [
        'Country name',
        'MF / LF / Form / CbCR',
        'What needs to be done',
        'Fixed / FYE-based / CIT-based / Upon Request',
        'YYYY-MM-DD if fixed',
        'Days / Months / Years',
        'Number (e.g., 30, 10, 12)',
        'Context and conditions'
    ]

    for col, instr in enumerate(deadline_instructions, start=1):
        cell = ws_deadlines.cell(row=2, column=col)
        cell.value = instr
        cell.fill = instruction_fill
        cell.font = Font(italic=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    deadline_examples = [
        ['Germany', 'MF', 'MF submission (if audit)', 'Upon Request', None, 'Days', 30, 'Within 30 days of audit notice'],
        ['Germany', 'LF', 'LF preparation', 'FYE-based', None, 'Months', 6, 'Within 6 months of FYE for extraordinary transactions'],
        ['Spain', 'Form', 'Form 232 filing', 'Fixed', '2026-08-25', None, None, 'Approximately 25 Aug each year'],
        ['Spain', 'MF', 'MF preparation', 'CIT-based', None, 'Days', -5, '5 days before CIT filing (approx 25 Jul)'],
    ]

    for row_idx, example in enumerate(deadline_examples, start=3):
        for col_idx, value in enumerate(example, start=1):
            cell = ws_deadlines.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.fill = example_fill

    ws_deadlines.column_dimensions['A'].width = 15
    ws_deadlines.column_dimensions['B'].width = 18
    ws_deadlines.column_dimensions['C'].width = 35
    ws_deadlines.column_dimensions['D'].width = 18
    ws_deadlines.column_dimensions['E'].width = 15
    ws_deadlines.column_dimensions['F'].width = 15
    ws_deadlines.column_dimensions['G'].width = 15
    ws_deadlines.column_dimensions['H'].width = 40

    ws_deadlines.row_dimensions[1].height = 30
    ws_deadlines.row_dimensions[2].height = 60

    # ===== SHEET 6: Instructions =====
    ws_instructions = wb.create_sheet("Instructions", 0)

    instructions_text = """
TP COMPLIANCE RULES LIBRARY - INSTRUCTIONS

This file defines the RULES for TP compliance requirements across jurisdictions.
It is SEPARATE from client data - you define these rules once, then reuse for all clients.

═══════════════════════════════════════════════════════════════════════════════

HOW THE RULE SYSTEM WORKS:

CONCEPT: Rules can have MULTIPLE CONDITIONS with AND/OR logic

Example: Germany Local File is required if:
  (Goods RPTs > EUR 6M) OR (Other RPTs > EUR 600K)

This is represented as TWO rows with the SAME Rule ID but DIFFERENT Condition Groups:

Rule ID    | Condition Group | Group Logic | Metric Type | Threshold
LF-DE-1   | 1               | OR          | RPTs (Goods)| 6000000
LF-DE-1   | 2               | OR          | RPTs (Other)| 600000

═══════════════════════════════════════════════════════════════════════════════

SHEET 1: MF REQUIREMENTS

Defines when Master File is required.

COLUMNS:
- Rule ID: Unique identifier (e.g., MF-DE-1, MF-ES-1)
- Country: Country name
- Condition Group: Number (1, 2, 3...) - conditions in SAME group use OR logic
- Group Logic: OR / AND between condition groups
- Metric Type: What to measure
    * Revenue: Total income/sales
    * Employees: Number of employees
    * Balance Sheet: Total assets/equity
    * RPTs: Related party transactions
    * Always: No threshold (always required)
- Metric Scope: What level
    * Group (Consolidated): Ultimate parent consolidated revenue
    * Local Entity: Just the local subsidiary
    * Transaction (Goods/Services/Other): Specific transaction types
- Threshold Value: Numeric threshold
- Currency: EUR, USD, GBP, etc.
- Operator: >= (greater than or equal), > (greater than), = (equal), etc.
- Notes: Additional context

EXAMPLE SCENARIOS:

Simple OR:
  Germany MF if (Group Revenue >= 750M) OR (Local Sales >= 100M)
  → Two rows, same Rule ID, different Condition Groups, both use OR

Complex AND with OR:
  Country X MF if (Group Revenue >= 500M) AND (Employees > 250 OR Balance Sheet > 43M)
  → Rule ID X-1: Condition Group 1, Revenue check
  → Rule ID X-1: Condition Group 2, Group Logic AND, Employees check
  → Rule ID X-1: Condition Group 3, Group Logic OR, Balance Sheet check

═══════════════════════════════════════════════════════════════════════════════

SHEET 2: LF REQUIREMENTS

Same structure as MF Requirements.
Defines when Local File is required.

Common metrics for LF:
- Local entity revenue/turnover
- RPT transaction amounts (Goods, Services, Financing, IP, Other)
- Number of RPT transactions
- Local entity employee count

═══════════════════════════════════════════════════════════════════════════════

SHEET 3: CBCR REQUIREMENTS

Usually simpler - typically just group revenue >= 750M EUR.
Same structure as MF/LF sheets.

═══════════════════════════════════════════════════════════════════════════════

SHEET 4: FORMS REQUIREMENTS

Defines which forms/disclosures are required.

COLUMNS:
- Country: Country name
- Form Name: Official form name (e.g., "Form 232", "Schedule 17-4")
- Form Type: TP Return / Disclosure / Notification / CbCR
- Condition Logic: When is it required?
    * Always: Every entity must file
    * If MF Required: Only if MF rules trigger
    * If LF Required: Only if LF rules trigger
    * Custom: Define specific rule
- Filing Deadline Rule: How to calculate deadline
- Notes: Additional context

═══════════════════════════════════════════════════════════════════════════════

SHEET 5: DEADLINES

Defines filing and preparation deadlines.

DEADLINE RULE TYPES:
1. Fixed: Same date every year (e.g., 2026-08-25)
2. FYE-based: Calculated from Fiscal Year End
   - Example: "FYE + 10 months" = 10 months after fiscal year end
3. CIT-based: Tied to Corporate Income Tax filing
   - Example: "CIT + 30 days" = 30 days after CIT return
4. Upon Request: Submitted when authorities request
   - Example: "30 days from request"

OFFSET TYPE: Days / Months / Years
OFFSET VALUE: Numeric (can be negative for "before")
  - Example: -5 days = 5 days BEFORE the reference date

═══════════════════════════════════════════════════════════════════════════════

FILLING IN THE RULES:

STEP 1: Start with examples (rows 3+) - they're real rules for Germany and Spain
STEP 2: Add more rows below for additional conditions
STEP 3: Use same Rule ID for related conditions
STEP 4: Don't delete instruction rows (row 2) - they help you remember format

METRIC TYPE OPTIONS:
- Revenue (most common)
- Employees
- Balance Sheet
- RPTs (specify type in Metric Scope)
- Assets
- Equity
- Transactions Count
- Always (no threshold)

METRIC SCOPE OPTIONS:
- Group (Consolidated)
- Local Entity
- Transaction (Goods)
- Transaction (Services)
- Transaction (Financing)
- Transaction (IP)
- Transaction (Other)
- Transaction (All)

OPERATOR OPTIONS:
>= (greater than or equal)
> (greater than)
= (equal to)
< (less than)
<= (less than or equal)

═══════════════════════════════════════════════════════════════════════════════

TIPS:

1. Most MF requirements use Group Revenue >= 750M EUR (OECD standard)
2. LF requirements vary widely - often local revenue + RPT thresholds
3. Use consistent Rule IDs: MF-[COUNTRY]-[NUMBER]
4. For complex conditions, sketch them out first:
   "If (A or B) and (C or D)"
   → Group 1: A (OR)
   → Group 1: B (OR)
   → Group 2: C (AND)
   → Group 2: D (OR)

5. Test your logic: Does it make sense? Can the system evaluate it?

═══════════════════════════════════════════════════════════════════════════════

NEXT STEP:

After filling this file, you'll create a CLIENT DATA file with:
- Group revenue
- Entity-level data (local revenue, RPTs, employees, etc.)

The system will:
1. Read these rules
2. Read client data
3. Apply rules to data
4. Determine what's required
5. Flag missing data
6. Generate dashboard

═══════════════════════════════════════════════════════════════════════════════

For questions or support, refer to the prototype documentation.

Version: 1.0 Prototype (Germany & Spain)
Last Updated: 2025-10-23
"""

    ws_instructions['A1'] = instructions_text
    ws_instructions['A1'].alignment = Alignment(wrap_text=True, vertical='top')
    ws_instructions['A1'].font = Font(name='Consolas', size=10)
    ws_instructions.column_dimensions['A'].width = 120
    ws_instructions.row_dimensions[1].height = 1400

    wb.save('Country_Rules_Library.xlsx')
    print("✓ Created: Country_Rules_Library.xlsx")


def create_client_data_template():
    """Create the Client Data input template"""

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Styling
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    instruction_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # ===== SHEET 1: Client Info =====
    ws_client = wb.create_sheet("Client Info")

    ws_client['A1'] = 'Field'
    ws_client['B1'] = 'Value'
    ws_client['A1'].fill = header_fill
    ws_client['B1'].fill = header_fill
    ws_client['A1'].font = header_font
    ws_client['B1'].font = header_font

    client_fields = [
        ['Client Name', 'Enter client name here'],
        ['Fiscal Year End (FYE)', '2025-12-31'],
        ['Fiscal Year Label', 'FYE 2025'],
        ['Group Revenue (EUR)', 'Enter total consolidated group revenue in EUR'],
        ['Group Revenue (Original Currency)', ''],
        ['Original Currency', 'e.g., USD, GBP, JPY'],
        ['Ultimate Holding Company', 'Name of UHC'],
        ['UHC Country', 'Country where UHC is located'],
        ['CbCR Filing Entity', 'Which entity files CbCR'],
        ['CbCR Filing Country', 'Country where CbCR is filed'],
        ['Data Completeness %', '0% (will be calculated)'],
    ]

    for row_idx, (field, value) in enumerate(client_fields, start=2):
        ws_client.cell(row=row_idx, column=1, value=field)
        ws_client.cell(row=row_idx, column=2, value=value)
        ws_client.cell(row=row_idx, column=2).fill = data_fill

    ws_client.column_dimensions['A'].width = 35
    ws_client.column_dimensions['B'].width = 50

    # ===== SHEET 2: Entity Data =====
    ws_entities = wb.create_sheet("Entity Data")

    entity_headers = [
        'Country',
        'Entity Name',
        'Local Revenue (EUR)',
        'Local Employees',
        'Balance Sheet Total (EUR)',
        'RPTs - Goods (EUR)',
        'RPTs - Services (EUR)',
        'RPTs - Financing (EUR)',
        'RPTs - IP (EUR)',
        'RPTs - Other (EUR)',
        'RPTs - Total (EUR)',
        'CIT Filing Date',
        'Data Complete?',
        'Missing Data'
    ]

    for col, header in enumerate(entity_headers, start=1):
        cell = ws_entities.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Instruction row
    entity_instructions = [
        'Germany or Spain',
        'Legal entity name',
        'Annual revenue in EUR',
        'Number of employees',
        'Total assets in EUR',
        'Goods RPT amounts',
        'Services RPT amounts',
        'Financing RPT amounts',
        'IP/Royalty RPT amounts',
        'Other RPT amounts',
        'Sum of all RPTs',
        'YYYY-MM-DD',
        'YES / NO / PARTIAL',
        'List what data is missing'
    ]

    for col, instr in enumerate(entity_instructions, start=1):
        cell = ws_entities.cell(row=2, column=col)
        cell.value = instr
        cell.fill = instruction_fill
        cell.font = Font(italic=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Example rows with question marks for unknown data
    example_entities = [
        ['Germany', 'Enter entity name', '?', '?', '?', '?', '?', '?', '?', '?', '?', '2026-07-30', 'NO', 'All financial data needed'],
        ['Spain', 'Enter entity name', '?', '?', '?', '?', '?', '?', '?', '?', '?', '2026-07-25', 'NO', 'All financial data needed'],
    ]

    for row_idx, example in enumerate(example_entities, start=3):
        for col_idx, value in enumerate(example, start=1):
            cell = ws_entities.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.fill = data_fill

    # Set column widths
    for col in range(1, 15):
        ws_entities.column_dimensions[get_column_letter(col)].width = 18

    ws_entities.column_dimensions['A'].width = 15
    ws_entities.column_dimensions['B'].width = 30
    ws_entities.column_dimensions['N'].width = 35

    ws_entities.row_dimensions[1].height = 40
    ws_entities.row_dimensions[2].height = 60

    # ===== SHEET 3: Instructions =====
    ws_instructions = wb.create_sheet("Instructions", 0)

    instructions_text = """
CLIENT DATA INPUT - INSTRUCTIONS

This file contains YOUR CLIENT'S SPECIFIC DATA.
It works with Country_Rules_Library.xlsx to determine compliance requirements.

═══════════════════════════════════════════════════════════════════════════════

HOW TO FILL THIS OUT:

SHEET 1: CLIENT INFO
Fill in high-level client information:
- Client Name
- Fiscal Year End (YYYY-MM-DD format)
- Group Revenue in EUR (consolidated revenue of ultimate parent)
- If revenue is in another currency, enter both original and converted EUR amounts

SHEET 2: ENTITY DATA
One row per entity (subsidiary) in scope.

For the PROTOTYPE, focus on Germany and Spain entities only.

REQUIRED FIELDS:
- Country: Must be "Germany" or "Spain"
- Entity Name: Legal name of the subsidiary

DATA FIELDS (use "?" if unknown):
- Local Revenue (EUR): Annual revenue of this entity
- Local Employees: Number of employees
- Balance Sheet Total (EUR): Total assets
- RPTs - [Type] (EUR): Related party transaction amounts by type
  * Goods: Purchase/sale of goods
  * Services: Management fees, technical services, etc.
  * Financing: Loans, interest, guarantees
  * IP: Royalties, license fees
  * Other: Everything else
  * Total: Sum of all RPTs

- CIT Filing Date: When is corporate tax return due (YYYY-MM-DD)
- Data Complete?:
  * YES = All data provided
  * NO = Significant data missing
  * PARTIAL = Some data provided
- Missing Data: Describe what's missing

═══════════════════════════════════════════════════════════════════════════════

USING "?" FOR UNKNOWN DATA:

If you don't know a value, enter "?" (question mark).

The system will:
1. Flag this as a data gap
2. Try to determine requirements with available data
3. Provide a "likely required" assessment
4. Generate a data request list for the client

EXAMPLE:
Germany entity with:
- Group revenue: EUR 850M (known)
- Local revenue: ? (unknown)
- RPTs: ? (unknown)

System output:
✓ MF: REQUIRED (group revenue > 750M threshold)
⚠️ LF: LIKELY REQUIRED (local revenue unknown but group is large)
📋 DATA NEEDED: Local revenue, Goods RPTs, Other RPTs

═══════════════════════════════════════════════════════════════════════════════

WHAT HAPPENS NEXT:

After you fill this file:

1. Save the file as: Client_Data_[ClientName].xlsx
2. Run: python apply_rules.py Country_Rules_Library.xlsx Client_Data_[ClientName].xlsx
3. System will:
   - Read country rules
   - Read your client data
   - Apply rules to determine requirements
   - Flag data gaps
   - Generate HTML dashboard with:
     * ✅ Confirmed requirements
     * ⚠️ Likely requirements (need verification)
     * 📋 Data request list for client
     * 🗓️ Filing deadlines

═══════════════════════════════════════════════════════════════════════════════

TIPS:

1. Enter data as accurately as possible
2. Use "?" for unknowns - don't guess
3. Convert all amounts to EUR for consistency
4. Include notes in "Missing Data" column
5. Update this file as you receive more information from client

═══════════════════════════════════════════════════════════════════════════════

PROTOTYPE SCOPE:

This prototype supports:
- Germany
- Spain
- Master File, Local File, CbCR, Forms requirements
- Multi-condition rules (e.g., revenue OR employees OR balance sheet)

Future versions will support all 22+ jurisdictions.

═══════════════════════════════════════════════════════════════════════════════

Version: 1.0 Prototype
Last Updated: 2025-10-23
"""

    ws_instructions['A1'] = instructions_text
    ws_instructions['A1'].alignment = Alignment(wrap_text=True, vertical='top')
    ws_instructions['A1'].font = Font(name='Consolas', size=10)
    ws_instructions.column_dimensions['A'].width = 120
    ws_instructions.row_dimensions[1].height = 1000

    wb.save('Client_Data_Template.xlsx')
    print("✓ Created: Client_Data_Template.xlsx")


if __name__ == "__main__":
    print("Creating Rule-Based TP Compliance Templates...\n")
    create_country_rules_library()
    create_client_data_template()
    print("\n✓ Templates created successfully!")
    print("\nNext steps:")
    print("1. Open Country_Rules_Library.xlsx - review/edit the rules for Germany and Spain")
    print("2. Open Client_Data_Template.xlsx - fill in your client's data")
    print("3. Run the rule engine to generate compliance assessment")
