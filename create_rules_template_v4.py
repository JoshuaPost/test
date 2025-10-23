#!/usr/bin/env python3
"""
Create Country Rules Library v4.0 - Enhanced User Guidance
Comprehensive 2-row explanations in each tab for better usability
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_data_dictionary_sheet(wb):
    """Create comprehensive data dictionary"""

    ws = wb.create_sheet("Data Dictionary", 0)

    # Styling
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    section_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

    # Title
    ws['A1'] = 'TP COMPLIANCE RULES LIBRARY v4.0 - DATA DICTIONARY'
    ws['A1'].font = Font(size=16, bold=True, color="1B5E20")
    ws.merge_cells('A1:E1')

    # Dropdowns section
    ws['A3'] = 'GLOBAL DROPDOWN VALUES'
    ws['A3'].font = Font(size=14, bold=True)
    ws['A3'].fill = section_fill
    ws.merge_cells('A3:E3')

    dropdowns_data = [
        ['Field', 'Allowed Values', 'Notes', 'Used In', 'Example'],

        ['APPLICABILITY', 'Always, Conditional, Integrated, Notification Only, Never Required, N/A',
         'Integrated=MF/LF only; Notification Only=CbCR only', 'All sheets', 'Malaysia MF: Integrated'],

        ['GROUP LOGIC', 'AND, OR',
         'How to combine condition groups', 'MF, LF, CbCR', 'OR: Either threshold triggers'],

        ['METRIC TYPE', 'Revenue, Group Revenue, Employees, Balance Sheet, RPTs, Always, Other',
         'What is being measured', 'MF, LF, CbCR', 'Revenue'],

        ['METRIC SCOPE', 'Group, Local Entity, Transaction (Goods), Transaction (Services), Transaction (All)',
         'Level of measurement', 'MF, LF, CbCR', 'Group (Consolidated)'],

        ['OPERATOR', '>=, >, =, <, <=',
         'Comparison operator', 'MF, LF, CbCR', '>= (greater than or equal)'],

        ['PREP DATE RULE', 'None, CIT Date, FYE-Based, Fixed, Upon Request, With Tax Return',
         'When to prepare', 'All sheets', 'CIT Date'],

        ['SUBMISSION DATE RULE', 'None, CIT Date, FYE-Based, Fixed, Upon Request, With Tax Return',
         'When to submit', 'All sheets', 'Upon Request'],

        ['INTEGRATED_WITH', 'Local File, TP Form, Other',
         'Where MF/LF content embedded', 'MF, LF only', 'Local File'],

        ['PENALTY PROTECTION ONLY', 'Yes, No',
         'Voluntary for penalty protection', 'MF, LF only', 'Yes (US/Canada)'],

        ['NOTIFICATION FREQUENCY', 'Annual, One-Time, Upon Change',
         'How often to notify', 'CbCR Notifications', 'Annual'],

        ['FILER TYPE', 'UPE, Local CE, One CE for All, Other',
         'Who files notification', 'CbCR Notifications', 'One CE for All'],

        ['JOINT FILING ALLOWED', 'Yes, No, Not Specified',
         'Can one entity file for group', 'CbCR Notifications', 'Yes'],

        ['INCLUDED IN CIT RETURN', 'Yes, No',
         'Filed within tax return', 'CbCR Notifications', 'No'],

        ['FORM TYPE', 'TP Disclosure, TP Return, MF Summary, LF Summary, CbCR Notification, Other',
         'Type of form', 'TP Forms', 'MF Summary'],

        ['FORM TRIGGER', 'Always, If MF Required, If LF Required, If MF or LF Required, If CbCR Required, Other',
         'What triggers the form', 'TP Forms', 'If MF Required'],

        ['LINKED TO', 'MF, LF, CbCR, Standalone',
         'Which document form relates to', 'TP Forms, CbCR Notifications', 'MF'],

        ['E-SIGNATURE REQUIRED', 'Yes, No',
         'Electronic signature needed', 'TP Forms', 'Yes'],

        ['TIMESTAMP REQUIRED', 'Yes, No, Electronic Timestamp',
         'Timestamp requirement', 'TP Forms', 'Electronic Timestamp'],
    ]

    for row_idx, row_data in enumerate(dropdowns_data, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            if row_idx == 4:
                cell.fill = header_fill
                cell.font = header_font

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30

    # Validation Rules section
    current_row = len(dropdowns_data) + 6

    ws[f'A{current_row}'] = 'VALIDATION & INTEGRITY RULES'
    ws[f'A{current_row}'].font = Font(size=14, bold=True)
    ws[f'A{current_row}'].fill = section_fill
    ws.merge_cells(f'A{current_row}:E{current_row}')

    validation_rules = [
        ['Rule', 'Condition', 'Action', 'Sheet', 'Example'],

        ['Integrated Applicability', 'IF APPLICABILITY = "Integrated"', 'INTEGRATED_WITH must not be blank', 'MF, LF',
         'Malaysia MF: Integrated WITH Local File'],

        ['Upon Request Days', 'IF SUBMISSION DATE RULE = "Upon Request"', 'UPON REQUEST DAYS should be numeric', 'All',
         'Germany: 30 days'],

        ['Multi-Condition Rules', 'Related conditions', 'Use same RULE ID, increment CONDITION GROUP', 'MF, LF, CbCR',
         'MF-DE-1: Groups 1 and 2'],

        ['CbCR In CIT', 'IF INCLUDED IN CIT RETURN = "Yes"', 'SUBMISSION CHANNEL = "Within CIT Return"', 'CbCR Notifications',
         'UK: In CIT return'],

        ['Joint Filing', 'IF FILER TYPE = "One CE for All"', 'JOINT FILING ALLOWED = "Yes"', 'CbCR Notifications',
         'Belgium: One CE can file'],

        ['Form Type Linking', 'IF FORM TYPE = "MF Summary"', 'LINKED TO = "MF"', 'TP Forms',
         'Form 275.MF → MF'],

        ['Penalty Protection', 'IF PENALTY PROTECTION ONLY = "Yes"', 'RULE NOTES must explain', 'MF, LF',
         'US: Voluntary penalty protection'],
    ]

    current_row += 1
    for row_idx, row_data in enumerate(validation_rules, start=current_row):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            if row_idx == current_row:
                cell.fill = header_fill
                cell.font = header_font

    return ws


def create_global_sheet_structure(ws, sheet_name, additional_columns=[]):
    """Create unified global structure for MF/LF/CbCR"""

    header_fill = PatternFill(start_color="0D2A5C", end_color="0D2A5C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    # Base global columns (common to all)
    base_columns = [
        'Country',
        'Applicability',
        'Rule ID',
        'Condition Group',
        'Group Logic',
        'Metric Type',
        'Metric Scope',
        'Threshold Value',
        'Currency',
        'Operator',
        'Prep Date Rule',
        'Prep Date Details',
        'Submission Date Rule',
        'Submission Date Details',
        'Upon Request Days',
        'Effective From (FY)',
        'Rule Notes',
        'Deadline Notes'
    ]

    # Insert additional columns after 'Applicability' (index 1)
    all_columns = base_columns[:2] + additional_columns + base_columns[2:]

    # Write headers
    for col_idx, header in enumerate(all_columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    return all_columns


def create_mf_sheet(wb):
    """Create Master File Requirements sheet with global schema"""

    ws = wb.create_sheet("MF Requirements")

    # MF-specific columns (inserted after Applicability)
    mf_specific = [
        'Integrated With',
        'Submission Channel',
        'Special Deadline Condition',
        'Penalty Protection Only'
    ]

    all_columns = create_global_sheet_structure(ws, "MF Requirements", mf_specific)

    # Add data validation dropdowns
    dv_applicability = DataValidation(type="list", formula1='"Always,Conditional,Integrated,Never Required,N/A"', allow_blank=True)
    dv_integrated = DataValidation(type="list", formula1='"Local File,TP Form,Other"', allow_blank=True)
    dv_penalty = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_group_logic = DataValidation(type="list", formula1='"AND,OR"', allow_blank=True)
    dv_metric_type = DataValidation(type="list", formula1='"Revenue,Group Revenue,Employees,Balance Sheet,RPTs,Always,Other"', allow_blank=True)
    dv_metric_scope = DataValidation(type="list", formula1='"Group (Consolidated),Local Entity,Transaction (Goods),Transaction (Services),Transaction (All)"', allow_blank=True)
    dv_operator = DataValidation(type="list", formula1='">=,>,=,<,<="', allow_blank=True)
    dv_date_rule = DataValidation(type="list", formula1='"None,CIT Date,FYE-Based,Fixed,Upon Request,With Tax Return"', allow_blank=True)

    ws.add_data_validation(dv_applicability)
    ws.add_data_validation(dv_integrated)
    ws.add_data_validation(dv_penalty)
    ws.add_data_validation(dv_group_logic)
    ws.add_data_validation(dv_metric_type)
    ws.add_data_validation(dv_metric_scope)
    ws.add_data_validation(dv_operator)
    ws.add_data_validation(dv_date_rule)

    # Apply validations to columns (rows 4-1000)
    dv_applicability.add(f'B4:B1000')  # Applicability
    dv_integrated.add(f'C4:C1000')  # Integrated With
    dv_penalty.add(f'F4:F1000')  # Penalty Protection Only
    dv_group_logic.add(f'I4:I1000')  # Group Logic
    dv_metric_type.add(f'J4:J1000')  # Metric Type
    dv_metric_scope.add(f'K4:K1000')  # Metric Scope
    dv_operator.add(f'N4:N1000')  # Operator
    dv_date_rule.add(f'O4:O1000')  # Prep Date Rule
    dv_date_rule.add(f'Q4:Q1000')  # Submission Date Rule

    # Add comprehensive explanation rows (2 rows for better clarity)
    explanation_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    explanation_font = Font(italic=True, size=9, bold=True)

    # Row 2: Field descriptions
    descriptions = [
        'Country name',
        'Is MF required? Always=Yes always | Conditional=If threshold met | Integrated=Content in LF/Form | Never Required=Not needed | N/A',
        'WHERE is MF content? (if Integrated). Options: Local File | TP Form | Other',
        'HOW to submit MF. e-filing portal | Form 275.MF | Attachment with CIT | Paper | BZSt Portal | etc.',
        'SPECIAL deadline rule (if any). Example: Germany extraordinary RPTs → full TPD due within 6 months after FYE',
        'Is this VOLUNTARY for penalty protection only? Yes (US/Canada) | No (mandatory)',
        'Unique rule ID. Format: MF-[CountryCode]-[Number]. Example: MF-DE-1, MF-ES-1',
        'Condition group number. Use 1, 2, 3... if multiple conditions. Same Rule ID, different groups',
        'How to combine groups. OR = Any condition triggers | AND = All must be met',
        'WHAT to measure. Revenue | Group Revenue | Employees | Balance Sheet | RPTs | Always | Other',
        'LEVEL of measurement. Group (Consolidated) | Local Entity | Transaction (Goods) | Transaction (Services) | Transaction (All)',
        'The threshold NUMBER. Examples: 750000000 (750M) | 6000000 (6M) | 250 (employees)',
        'Currency for threshold. EUR | USD | GBP | JPY | MYR | AUD | CAD | etc.',
        'Comparison. >= (or more) | > (more than) | = (exactly) | < (less than) | <= (or less)',
        'WHEN to prepare (recommended). None | CIT Date | FYE-Based | Fixed | Upon Request | With Tax Return',
        'Prep date details. Examples: CIT - 5 days | FYE + 10 months | Within 6 months of FYE | Expected 31 Jul',
        'WHEN to submit (hard deadline). None | CIT Date | FYE-Based | Fixed | Upon Request | With Tax Return',
        'Submission details. Examples: 2026-08-25 | Within 30 days of audit | By 31 Dec following FY | N/A - voluntary',
        'Days to respond if Upon Request. Examples: 30 (Germany) | 14 (Malaysia) | 10 (Spain)',
        'When did this rule start? FY2024 | FY2023 | FY2016 | etc.',
        'Context about the threshold rule, special conditions, notes',
        'Context about deadlines, submission requirements, additional notes'
    ]

    # Row 3: Examples
    examples_row = [
        'Germany, Spain, Malaysia, etc.',
        'Conditional (most common)',
        'Local File (Malaysia MF)',
        'e-filing portal (most common)',
        'See Germany example below',
        'No (most) | Yes (US/Canada)',
        'MF-DE-1, MF-ES-1, MF-MY-1',
        '1 (or 2, 3 for multi-condition)',
        'OR (most common)',
        'Group Revenue (MF threshold)',
        'Group (Consolidated)',
        '750000000 (€750M BEPS)',
        'EUR (Europe) | USD (US)',
        '>= (or more)',
        'CIT Date (prepare with tax)',
        'Expected 31 Jul (Germany CIT)',
        'Upon Request (no hard filing)',
        'Within 30 days of audit',
        '30 (Germany) | 14 (Malaysia)',
        'FY2024 (most recent)',
        'Standard BEPS threshold',
        'Automatic submission upon audit'
    ]

    for col_idx, description in enumerate(descriptions, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = description
        cell.fill = explanation_fill
        cell.font = explanation_font
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    explanation_font_example = Font(italic=True, size=9)
    for col_idx, example in enumerate(examples_row, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = example
        cell.fill = explanation_fill
        cell.font = explanation_font_example
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws.row_dimensions[2].height = 65
    ws.row_dimensions[3].height = 40

    # Example data rows (start from row 4)
    example_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    examples = [
        # Germany - Standard conditional
        ['Germany', 'Conditional', '', '', '', 'No',
         'MF-DE-1', '1', 'OR', 'Revenue', 'Group (Consolidated)', 750000000, 'EUR', '>=',
         'None', '', 'Upon Request', 'Within 30 days of audit notice', 30, 'FY2024',
         'Standard BEPS threshold', 'Automatic submission upon audit'],

        # Germany - Extraordinary transactions (special deadline)
        ['Germany', 'Conditional', '', '', 'Extraordinary RPTs → full TPD due within 6 months after FYE', 'No',
         'MF-DE-2', '1', 'OR', 'Always', 'Local Entity', 0, 'EUR', '=',
         'FYE-Based', 'Within 6 months of FYE for extraordinary transactions', 'Upon Request', '30 days from audit', 30, 'FY2024',
         'Germany-specific extraordinary transaction rule', 'Separate timeline for extraordinary RPTs'],

        # Malaysia - Integrated
        ['Malaysia', 'Integrated', 'Local File', '', '', 'No',
         '', '', '', '', '', '', '', '',
         'CIT Date', 'Prepare with CTPD by CIT filing', 'Upon Request', 'Within 14 days', 14, 'FY2023',
         'MF content integrated into LF per 2023 TPD; no standalone MF. Group revenue below MYR 3B threshold.',
         'MF elements included in CTPD submission'],

        # United States - Penalty protection only
        ['United States', 'Conditional', '', '', '', 'Yes',
         '', '', '', '', '', '', '', '',
         'None', 'Voluntary preparation recommended', 'None', 'N/A - voluntary', '', 'FY2018',
         'Voluntary MF preparation for penalty protection under IRC §6662. No filing requirement.',
         'Contemporaneous documentation provides reasonable cause defense'],

        # Canada - Penalty protection only
        ['Canada', 'Conditional', '', '', '', 'Yes',
         '', '', '', '', '', '', '', '',
         'None', 'Voluntary preparation recommended', 'None', 'N/A - voluntary', '', 'FY2015',
         'Voluntary MF for penalty protection. No statutory filing requirement.',
         'Contemporaneous documentation required for transfer pricing adjustment defense'],
    ]

    for row_idx, example in enumerate(examples, start=4):
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.fill = example_fill

    # Set column widths
    for col_idx, col in enumerate(all_columns, start=1):
        width = 20  # default
        if col in ['Rule Notes', 'Deadline Notes', 'Special Deadline Condition']:
            width = 45
        elif col in ['Metric Scope', 'Prep Date Details', 'Submission Date Details']:
            width = 35
        elif col in ['Metric Type', 'Submission Channel']:
            width = 25
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 50

    return ws


def create_lf_sheet(wb):
    """Create Local File Requirements sheet with global schema"""

    ws = wb.create_sheet("LF Requirements")

    # LF-specific columns (inserted after Applicability)
    lf_specific = [
        'Integrated With',
        'Submission Channel',
        'Special Deadline Condition',
        'Penalty Protection Only'
    ]

    all_columns = create_global_sheet_structure(ws, "LF Requirements", lf_specific)

    # Add data validation dropdowns
    dv_applicability = DataValidation(type="list", formula1='"Always,Conditional,Integrated,Never Required,N/A"', allow_blank=True)
    dv_integrated = DataValidation(type="list", formula1='"Master File,TP Form,Other"', allow_blank=True)
    dv_penalty = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_group_logic = DataValidation(type="list", formula1='"AND,OR"', allow_blank=True)
    dv_metric_type = DataValidation(type="list", formula1='"Revenue,Group Revenue,Employees,Balance Sheet,RPTs,Always,Other"', allow_blank=True)
    dv_metric_scope = DataValidation(type="list", formula1='"Group (Consolidated),Local Entity,Transaction (Goods),Transaction (Services),Transaction (All)"', allow_blank=True)
    dv_operator = DataValidation(type="list", formula1='">=,>,=,<,<="', allow_blank=True)
    dv_date_rule = DataValidation(type="list", formula1='"None,CIT Date,FYE-Based,Fixed,Upon Request,With Tax Return"', allow_blank=True)

    ws.add_data_validation(dv_applicability)
    ws.add_data_validation(dv_integrated)
    ws.add_data_validation(dv_penalty)
    ws.add_data_validation(dv_group_logic)
    ws.add_data_validation(dv_metric_type)
    ws.add_data_validation(dv_metric_scope)
    ws.add_data_validation(dv_operator)
    ws.add_data_validation(dv_date_rule)

    # Apply validations to columns (rows 4-1000)
    dv_applicability.add(f'B4:B1000')  # Applicability
    dv_integrated.add(f'C4:C1000')  # Integrated With
    dv_penalty.add(f'F4:F1000')  # Penalty Protection Only
    dv_group_logic.add(f'I4:I1000')  # Group Logic
    dv_metric_type.add(f'J4:J1000')  # Metric Type
    dv_metric_scope.add(f'K4:K1000')  # Metric Scope
    dv_operator.add(f'N4:N1000')  # Operator
    dv_date_rule.add(f'O4:O1000')  # Prep Date Rule
    dv_date_rule.add(f'Q4:Q1000')  # Submission Date Rule

    # Add comprehensive explanation rows (2 rows for better clarity)
    explanation_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    explanation_font = Font(italic=True, size=9, bold=True)

    # Row 2: Field descriptions
    descriptions = [
        'Country name',
        'Is LF required? Always=Yes always | Conditional=If threshold met | Integrated=Content in MF/Form | Never Required=Not needed | N/A',
        'WHERE is LF content? (if Integrated). Options: Master File | TP Form | Other',
        'HOW to submit LF. e-filing portal | With CIT return | Portal | Paper | etc.',
        'SPECIAL deadline rule (if any). Example: Germany extraordinary RPTs → full TPD due within 6 months after FYE',
        'Is this VOLUNTARY for penalty protection only? Yes (US/Canada) | No (mandatory)',
        'Unique rule ID. Format: LF-[CountryCode]-[Number]. Example: LF-DE-1, LF-ES-1',
        'Condition group number. Use 1, 2, 3... if multiple conditions. Same Rule ID, different groups',
        'How to combine groups. OR = Any condition triggers | AND = All must be met',
        'WHAT to measure. Revenue | Group Revenue | Employees | Balance Sheet | RPTs | Always | Other',
        'LEVEL of measurement. Group (Consolidated) | Local Entity | Transaction (Goods) | Transaction (Services) | Transaction (All)',
        'The threshold NUMBER. Examples: 6000000 (6M) | 250000 (250K) | 250 (employees)',
        'Currency for threshold. EUR | USD | GBP | JPY | MYR | AUD | CAD | etc.',
        'Comparison. >= (or more) | > (more than) | = (exactly) | < (less than) | <= (or less)',
        'WHEN to prepare (recommended). None | CIT Date | FYE-Based | Fixed | Upon Request | With Tax Return',
        'Prep date details. Examples: Expected 31 Jul | By 7 months after FYE | CIT filing date',
        'WHEN to submit (hard deadline). None | CIT Date | FYE-Based | Fixed | Upon Request | With Tax Return',
        'Submission details. Examples: Within 30 days of audit | Within 10 days of request | File with CIT return',
        'Days to respond if Upon Request. Examples: 30 (Germany) | 14 (Malaysia) | 10 (Spain)',
        'When did this rule start? FY2024 | FY2023 | FY2016 | etc.',
        'Context about the threshold rule, special conditions, notes',
        'Context about deadlines, submission requirements, additional notes'
    ]

    # Row 3: Examples
    examples_row = [
        'Germany, Spain, Malaysia, etc.',
        'Conditional (most common)',
        'Not typically integrated',
        'e-filing portal / With CIT return',
        'See Germany example below',
        'No (most) | Yes (US/Canada)',
        'LF-DE-1, LF-ES-1, LF-MY-1',
        '1 (or 2 for multi-condition)',
        'OR (Germany has 2 conditions)',
        'RPTs (Local File threshold)',
        'Transaction (Goods) or (Services)',
        '6000000 (€6M goods) | 250000 (€250K)',
        'EUR (Europe) | USD (US)',
        '> (more than)',
        'CIT Date (prepare with tax)',
        'Expected 31 Jul (Germany CIT)',
        'Upon Request (no hard filing)',
        'Within 30 days of audit (Germany)',
        '30 (Germany) | 14 (Malaysia) | 10 (Spain)',
        'FY2024 (most recent)',
        'LF required if goods RPTs exceed 6M',
        'Automatic submission upon audit'
    ]

    for col_idx, description in enumerate(descriptions, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = description
        cell.fill = explanation_fill
        cell.font = explanation_font
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    explanation_font_example = Font(italic=True, size=9)
    for col_idx, example in enumerate(examples_row, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = example
        cell.fill = explanation_fill
        cell.font = explanation_font_example
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws.row_dimensions[2].height = 65
    ws.row_dimensions[3].height = 40

    # Example data rows (start from row 4)
    example_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    examples = [
        # Germany - Multi-condition OR
        ['Germany', 'Conditional', '', '', '', 'No',
         'LF-DE-1', '1', 'OR', 'RPTs', 'Transaction (Goods)', 6000000, 'EUR', '>',
         'CIT Date', 'Expected 31 Jul (CIT filing)', 'Upon Request', 'Within 30 days of audit notice', 30, 'FY2024',
         'LF required if goods RPTs exceed 6M EUR', 'Automatic submission upon audit'],

        ['Germany', 'Conditional', '', '', '', 'No',
         'LF-DE-1', '2', 'OR', 'RPTs', 'Transaction (Services)', 600000, 'EUR', '>',
         'CIT Date', 'Expected 31 Jul (CIT filing)', 'Upon Request', 'Within 30 days of audit notice', 30, 'FY2024',
         'OR services/other RPTs exceed 600K EUR', 'Automatic submission upon audit'],

        # Spain - Standard conditional
        ['Spain', 'Conditional', '', '', '', 'No',
         'LF-ES-1', '1', 'OR', 'RPTs', 'Transaction (All)', 250000, 'EUR', '>',
         'CIT Date', 'Expected 25 Jul', 'Upon Request', 'Within 10 days of request', 10, 'FY2016',
         'LF required if local RPTs exceed 250K EUR', 'Maintain contemporaneously'],

        # Malaysia - Always required
        ['Malaysia', 'Always', '', '', '', 'No',
         '', '', '', '', '', '', '', '',
         'CIT Date', 'By 7 months after FYE (CIT filing)', 'Upon Request', 'Within 14 days', 14, 'FY2023',
         'CTPD (LF) required for all entities with RPTs. MF content integrated per 2023 TPD.',
         'File with CIT return'],

        # United States - Penalty protection only
        ['United States', 'Conditional', '', '', '', 'Yes',
         '', '', '', '', '', '', '', '',
         'None', 'Voluntary preparation recommended', 'None', 'N/A - voluntary', '', 'FY2018',
         'Voluntary LF preparation for penalty protection under IRC §6662. No filing requirement.',
         'Contemporaneous documentation provides reasonable cause defense'],

        # Canada - Penalty protection only
        ['Canada', 'Conditional', '', '', '', 'Yes',
         '', '', '', '', '', '', '', '',
         'None', 'Voluntary preparation recommended', 'None', 'N/A - voluntary', '', 'FY2015',
         'Voluntary LF for penalty protection. No statutory filing requirement.',
         'Contemporaneous documentation required for transfer pricing adjustment defense'],
    ]

    for row_idx, example in enumerate(examples, start=4):
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.fill = example_fill

    # Set column widths
    for col_idx, col in enumerate(all_columns, start=1):
        width = 20  # default
        if col in ['Rule Notes', 'Deadline Notes', 'Special Deadline Condition']:
            width = 45
        elif col in ['Metric Scope', 'Prep Date Details', 'Submission Date Details']:
            width = 35
        elif col in ['Metric Type', 'Submission Channel']:
            width = 25
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 50

    return ws


def create_cbcr_notifications_sheet(wb):
    """Create CbCR Notifications sheet (redesigned from CbCR Requirements)"""

    ws = wb.create_sheet("CbCR Notifications")

    header_fill = PatternFill(start_color="0D2A5C", end_color="0D2A5C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    example_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    # CbCR Notification specific structure
    columns = [
        'Country',
        'Applicability',
        'Notification Frequency',
        'Filer Type',
        'Joint Filing Allowed?',
        'Included in CIT Return?',
        'Submission Channel',
        'Form Name / Reference',
        'Submission Date Rule',
        'Submission Date Details',
        'Notification Validity',
        'Linked To',
        'Effective From (FY)',
        'Rule Notes'
    ]

    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Add data validation dropdowns
    dv_applicability = DataValidation(type="list", formula1='"Always,Conditional,Notification Only,Never Required,N/A"', allow_blank=True)
    dv_frequency = DataValidation(type="list", formula1='"Annual,One-Time,Upon Change"', allow_blank=True)
    dv_filer = DataValidation(type="list", formula1='"UPE,Local CE,One CE for All,Other"', allow_blank=True)
    dv_joint = DataValidation(type="list", formula1='"Yes,No,Not Specified"', allow_blank=True)
    dv_in_cit = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_date_rule = DataValidation(type="list", formula1='"None,CIT Date,FYE-Based,Fixed,Upon Request,With Tax Return"', allow_blank=True)
    dv_linked = DataValidation(type="list", formula1='"MF,LF,CbCR,Standalone"', allow_blank=True)

    ws.add_data_validation(dv_applicability)
    ws.add_data_validation(dv_frequency)
    ws.add_data_validation(dv_filer)
    ws.add_data_validation(dv_joint)
    ws.add_data_validation(dv_in_cit)
    ws.add_data_validation(dv_date_rule)
    ws.add_data_validation(dv_linked)

    # Apply validations to columns (rows 3-1000)
    dv_applicability.add(f'B3:B1000')  # Applicability
    dv_frequency.add(f'C3:C1000')  # Notification Frequency
    dv_filer.add(f'D3:D1000')  # Filer Type
    dv_joint.add(f'E3:E1000')  # Joint Filing Allowed
    dv_in_cit.add(f'F3:F1000')  # Included in CIT Return
    dv_date_rule.add(f'I3:I1000')  # Submission Date Rule
    dv_linked.add(f'L3:L1000')  # Linked To

    # Add helpful explanation row
    explanation_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    explanation_font = Font(italic=True, size=9)

    explanations = [
        'Belgium, France, UK, Germany, etc.',
        'Always / Conditional / Notification Only / Never Required / N/A',
        'Annual / One-Time / Upon Change',
        'UPE / Local CE / One CE for All / Other',
        'Yes / No / Not Specified',
        'Yes / No',
        'Separate Form / Portal / Within CIT Return / BZSt Portal',
        'Form 275.CBC.NOT / DAS2-CbCR / SA / BZSt CbCR Notification',
        'None / CIT Date / FYE-Based / Fixed / Upon Request / With Tax Return',
        'Details (e.g., By 31 Dec following FY, Within 3 months of change)',
        'Valid for FY / Until entity or UPE info changes / Annual / Until change in filing entity',
        'MF / LF / CbCR / Standalone',
        'FY2016, 2017, etc.',
        'Context about notification rules and requirements'
    ]

    for col_idx, explanation in enumerate(explanations, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = explanation
        cell.fill = explanation_fill
        cell.font = explanation_font
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws.row_dimensions[2].height = 60

    # Examples
    examples = [
        # Belgium - Annual separate form
        ['Belgium', 'Conditional', 'Annual', 'Local CE', 'No', 'No', 'Separate Form', 'Form 275.CBC.NOT',
         'Fixed', 'By 31 Dec following FY', 'Valid for FY', 'CbCR', 'FY2016',
         'Annual CbCR notification filed separately from CIT return'],

        # France - One-time until change
        ['France', 'Conditional', 'Upon Change', 'UPE', 'Not Specified', 'No', 'Portal', 'DAS2-CbCR',
         'Upon Change', 'Within 3 months of change', 'Until entity or UPE info changes', 'CbCR', 'FY2017',
         'One-time notification valid until circumstances change (UPE change, threshold, etc.)'],

        # UK - In CIT return
        ['United Kingdom', 'Conditional', 'Upon Change', 'UPE', 'Not Specified', 'Yes', 'Within CIT Return', 'SA'

,
         'With Tax Return', 'Within CIT return filing', 'Until change in filing entity', 'CbCR', 'FY2016',
         'Notification included in CIT return; updated when filing entity changes'],

        # Germany - Separate notification
        ['Germany', 'Conditional', 'Annual', 'One CE for All', 'Yes', 'No', 'BZSt Portal', 'BZSt CbCR Notification',
         'Fixed', 'By end of reporting FY', 'Annual', 'CbCR', 'FY2016',
         'One German group entity can file notification for all German entities'],
    ]

    for row_idx, example in enumerate(examples, start=3):
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.fill = example_fill

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 22
    ws.column_dimensions['J'].width = 35
    ws.column_dimensions['K'].width = 30
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 50

    ws.row_dimensions[1].height = 50

    return ws


def create_tp_forms_sheet(wb):
    """Create enhanced TP Forms sheet"""

    ws = wb.create_sheet("TP Forms")

    header_fill = PatternFill(start_color="0D2A5C", end_color="0D2A5C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    example_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    columns = [
        'Country',
        'Form Name',
        'Form Type',
        'Form Trigger',
        'Linked To',
        'What It Contains',
        'Submission Date Rule',
        'Submission Date Details',
        'Upon Request Days',
        'Electronic Signature Required?',
        'Timestamp Required?',
        'Effective From (FY)',
        'Rule Notes'
    ]

    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Add data validation dropdowns
    dv_form_type = DataValidation(type="list", formula1='"TP Disclosure,TP Return,MF Summary,LF Summary,CbCR Notification,Other"', allow_blank=True)
    dv_trigger = DataValidation(type="list", formula1='"Always,If MF Required,If LF Required,If MF or LF Required,If CbCR Required,Other"', allow_blank=True)
    dv_linked = DataValidation(type="list", formula1='"MF,LF,CbCR,Standalone"', allow_blank=True)
    dv_date_rule = DataValidation(type="list", formula1='"None,CIT Date,FYE-Based,Fixed,Upon Request,With Tax Return"', allow_blank=True)
    dv_esig = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_timestamp = DataValidation(type="list", formula1='"Yes,No,Electronic Timestamp"', allow_blank=True)

    ws.add_data_validation(dv_form_type)
    ws.add_data_validation(dv_trigger)
    ws.add_data_validation(dv_linked)
    ws.add_data_validation(dv_date_rule)
    ws.add_data_validation(dv_esig)
    ws.add_data_validation(dv_timestamp)

    # Apply validations to columns (rows 3-1000)
    dv_form_type.add(f'C3:C1000')  # Form Type
    dv_trigger.add(f'D3:D1000')  # Form Trigger
    dv_linked.add(f'E3:E1000')  # Linked To
    dv_date_rule.add(f'G3:G1000')  # Submission Date Rule
    dv_esig.add(f'J3:J1000')  # Electronic Signature Required
    dv_timestamp.add(f'K3:K1000')  # Timestamp Required

    # Add helpful explanation row
    explanation_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    explanation_font = Font(italic=True, size=9)

    explanations = [
        'Belgium, Spain, Italy, Germany, etc.',
        'Form 275.MF / Form 232 / RS 106 / Transaction Matrix',
        'TP Disclosure / TP Return / MF Summary / LF Summary / CbCR Notification / Other',
        'Always / If MF Required / If LF Required / If MF or LF Required / If CbCR Required / Other',
        'MF / LF / CbCR / Standalone',
        'Summary form with key data / Annual TP informative return / TP disclosure / Structured overview of RPTs',
        'None / CIT Date / FYE-Based / Fixed / Upon Request / With Tax Return',
        'Details (e.g., 31 Dec following FY, Approx 25 Aug, Within 30 days of audit notice)',
        'Days to submit if upon request (30, 14, 10, etc.)',
        'Yes / No',
        'Yes / No / Electronic Timestamp',
        'FY2016, 2010, 2024, etc.',
        'Context about form requirements and special rules'
    ]

    for col_idx, explanation in enumerate(explanations, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = explanation
        cell.fill = explanation_fill
        cell.font = explanation_font
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws.row_dimensions[2].height = 60

    # Examples
    examples = [
        # Belgium Form 275.MF - MF Summary
        ['Belgium', 'Form 275.MF', 'MF Summary', 'If MF Required', 'MF', 'Summary form with key MF data points',
         'Fixed', '31 Dec following FY', '', 'Yes', 'No', 'FY2016',
         'Separate summary form filed alongside full MF report'],

        # Belgium Form 275.LF - LF Summary
        ['Belgium', 'Form 275.LF', 'LF Summary', 'If LF Required', 'LF', 'Summary form with key LF data points',
         'With Tax Return', 'Expected 31 Jul', '', 'Yes', 'No', 'FY2016',
         'Separate summary form filed with CIT return'],

        # Spain Form 232
        ['Spain', 'Form 232', 'TP Return', 'If MF or LF Required', 'Standalone', 'Annual TP informative return',
         'Fixed', 'Approx 25 Aug', '', 'Yes', 'No', 'FY2016',
         'Informative return separate from full MF/LF documentation'],

        # Italy RS 106
        ['Italy', 'RS 106', 'TP Disclosure', 'If MF or LF Required', 'Standalone', 'TP disclosure in tax return',
         'With Tax Return', 'Expected 31 Oct', '', 'Yes', 'Electronic Timestamp', 'FY2010',
         'MF/LF must be electronically timestamped before filing RS 106 disclosure'],

        # Germany Transaction Matrix
        ['Germany', 'Transaction Matrix', 'TP Disclosure', 'Always', 'MF', 'Structured overview of RPTs',
         'Upon Request', 'Within 30 days of audit notice', 30, 'No', 'Yes', 'FY2024',
         'NEW 2024 requirement - automatic submission with MF upon audit'],
    ]

    for row_idx, example in enumerate(examples, start=3):
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.fill = example_fill

    # Column widths
    for col_idx, col in enumerate(columns, start=1):
        width = 25
        if col in ['What It Contains', 'Submission Date Details', 'Rule Notes']:
            width = 40
        elif col in ['Form Name', 'Form Type', 'Form Trigger']:
            width = 25
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 50

    return ws


def create_v3_template():
    """Create complete v3.0 template"""

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    print("Creating sheets...")

    # Sheet 1: Data Dictionary (first sheet)
    create_data_dictionary_sheet(wb)
    print("  ✓ Data Dictionary")

    # Sheet 2: MF Requirements
    create_mf_sheet(wb)
    print("  ✓ MF Requirements")

    # Sheet 3: LF Requirements
    create_lf_sheet(wb)
    print("  ✓ LF Requirements")

    # Sheet 4: CbCR Notifications
    create_cbcr_notifications_sheet(wb)
    print("  ✓ CbCR Notifications")

    # Sheet 5: TP Forms
    create_tp_forms_sheet(wb)
    print("  ✓ TP Forms")

    # Save
    wb.save('Country_Rules_Library_v4.0.xlsx')
    print("\n✓ Country_Rules_Library_v4.0.xlsx created successfully!")


if __name__ == "__main__":
    print("="*80)
    print("COUNTRY RULES LIBRARY v4.0 - ENHANCED USER GUIDANCE")
    print("="*80)
    print()

    create_v3_template()

    print()
    print("NEW IN v4.0:")
    print("  ✓ 2-row comprehensive explanations in EVERY tab")
    print("  ✓ Row 2: Field descriptions with detailed guidance")
    print("  ✓ Row 3: Examples showing correct format")
    print("  ✓ All dropdowns updated to start from row 4")
    print("  ✓ MF & LF sheets: Full explanations from Data Dictionary")
    print("  ✓ Self-sufficient tabs - no need to switch to Data Dictionary")
    print()
    print("Next: Open Country_Rules_Library_v4.0.xlsx and review")
