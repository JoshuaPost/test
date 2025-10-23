#!/usr/bin/env python3
"""
TP Compliance Rule Engine - Prototype
Reads country rules and client data, applies logic, generates assessment
"""

import openpyxl
import sys
from collections import defaultdict

def read_rules(rules_file):
    """Read all rules from Country Rules Library"""

    wb = openpyxl.load_workbook(rules_file)

    rules = {
        'mf': [],
        'lf': [],
        'cbcr': [],
        'forms': [],
        'deadlines': []
    }

    # Read MF Requirements
    ws = wb['MF Requirements']
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:  # Skip empty rows
            continue
        rules['mf'].append({
            'rule_id': row[0],
            'country': row[1],
            'condition_group': row[2],
            'group_logic': row[3],
            'metric_type': row[4],
            'metric_scope': row[5],
            'threshold': row[6],
            'currency': row[7],
            'operator': row[8],
            'notes': row[9]
        })

    # Read LF Requirements
    ws = wb['LF Requirements']
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        rules['lf'].append({
            'rule_id': row[0],
            'country': row[1],
            'condition_group': row[2],
            'group_logic': row[3],
            'metric_type': row[4],
            'metric_scope': row[5],
            'threshold': row[6],
            'currency': row[7],
            'operator': row[8],
            'notes': row[9]
        })

    # Read CbCR Requirements
    ws = wb['CbCR Requirements']
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        rules['cbcr'].append({
            'rule_id': row[0],
            'country': row[1],
            'condition_group': row[2],
            'group_logic': row[3],
            'metric_type': row[4],
            'metric_scope': row[5],
            'threshold': row[6],
            'currency': row[7],
            'operator': row[8],
            'notes': row[9]
        })

    # Read Forms
    ws = wb['Forms Requirements']
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        rules['forms'].append({
            'country': row[0],
            'form_name': row[1],
            'form_type': row[2],
            'condition_logic': row[3],
            'deadline_rule': row[4],
            'notes': row[5]
        })

    # Read Deadlines
    ws = wb['Deadlines']
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        rules['deadlines'].append({
            'country': row[0],
            'requirement_type': row[1],
            'description': row[2],
            'deadline_rule': row[3],
            'fixed_date': row[4],
            'offset_type': row[5],
            'offset_value': row[6],
            'notes': row[7]
        })

    return rules


def read_client_data(client_file):
    """Read client data from template"""

    wb = openpyxl.load_workbook(client_file)

    # Read Client Info
    ws = wb['Client Info']
    client_info = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            client_info[row[0]] = row[1]

    # Read Entity Data
    ws = wb['Entity Data']
    entities = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        entities.append({
            'country': row[0],
            'entity_name': row[1],
            'local_revenue': row[2],
            'local_employees': row[3],
            'balance_sheet': row[4],
            'rpts_goods': row[5],
            'rpts_services': row[6],
            'rpts_financing': row[7],
            'rpts_ip': row[8],
            'rpts_other': row[9],
            'rpts_total': row[10],
            'cit_filing_date': row[11],
            'data_complete': row[12],
            'missing_data': row[13]
        })

    return {
        'client_info': client_info,
        'entities': entities
    }


def evaluate_condition(condition, client_data, entity_data):
    """Evaluate a single condition against client/entity data"""

    metric_type = condition['metric_type']
    metric_scope = condition['metric_scope']
    threshold = condition['threshold']
    operator = condition['operator']

    # Get the value to compare
    value = None
    data_available = True

    if 'Group' in str(metric_scope):
        # Group-level metric
        if 'Revenue' in metric_type:
            group_rev = client_data['client_info'].get('Group Revenue (EUR)')
            if group_rev and group_rev != '?' and not isinstance(group_rev, str):
                value = float(group_rev)
            elif group_rev == '?' or not group_rev:
                data_available = False

    elif 'Local Entity' in str(metric_scope):
        # Entity-level metric
        if 'Revenue' in metric_type or 'Turnover' in metric_type:
            local_rev = entity_data.get('local_revenue')
            if local_rev and local_rev != '?':
                try:
                    value = float(local_rev)
                except (ValueError, TypeError):
                    data_available = False
            else:
                data_available = False

        elif 'Employees' in metric_type:
            employees = entity_data.get('local_employees')
            if employees and employees != '?':
                try:
                    value = float(employees)
                except (ValueError, TypeError):
                    data_available = False
            else:
                data_available = False

        elif 'Balance Sheet' in metric_type or 'Assets' in metric_type:
            bs = entity_data.get('balance_sheet')
            if bs and bs != '?':
                try:
                    value = float(bs)
                except (ValueError, TypeError):
                    data_available = False
            else:
                data_available = False

    elif 'Transaction' in str(metric_scope):
        # RPT metrics
        if 'Goods' in str(metric_scope):
            rpt = entity_data.get('rpts_goods')
        elif 'Services' in str(metric_scope):
            rpt = entity_data.get('rpts_services')
        elif 'Financing' in str(metric_scope):
            rpt = entity_data.get('rpts_financing')
        elif 'IP' in str(metric_scope):
            rpt = entity_data.get('rpts_ip')
        elif 'Other' in str(metric_scope):
            rpt = entity_data.get('rpts_other')
        elif 'All' in str(metric_scope):
            rpt = entity_data.get('rpts_total')
        else:
            rpt = None

        if rpt and rpt != '?':
            try:
                value = float(rpt)
            except (ValueError, TypeError):
                data_available = False
        else:
            data_available = False

    # If data not available, return unknown
    if not data_available or value is None:
        return {
            'result': 'UNKNOWN',
            'reason': f"{metric_type} ({metric_scope}) data not provided",
            'metric': metric_type,
            'scope': metric_scope
        }

    # Evaluate condition
    result = False
    if operator == '>=':
        result = value >= threshold
    elif operator == '>':
        result = value > threshold
    elif operator == '=':
        result = value == threshold
    elif operator == '<':
        result = value < threshold
    elif operator == '<=':
        result = value <= threshold

    return {
        'result': 'TRUE' if result else 'FALSE',
        'reason': f"{metric_type} ({value:,.0f}) {operator} {threshold:,.0f}",
        'metric': metric_type,
        'scope': metric_scope,
        'value': value,
        'threshold': threshold
    }


def evaluate_rule_set(rule_set, client_data, entity_data):
    """Evaluate a complete rule set with AND/OR logic"""

    if not rule_set:
        return {'required': False, 'confidence': 'N/A', 'details': []}

    # Group rules by rule_id
    rules_by_id = defaultdict(list)
    for rule in rule_set:
        rules_by_id[rule['rule_id']].append(rule)

    # Evaluate each rule ID
    results = []
    for rule_id, rules in rules_by_id.items():
        # Group by condition group
        groups = defaultdict(list)
        for rule in rules:
            groups[rule['condition_group']].append(rule)

        # Evaluate each group (conditions within group use OR)
        group_results = []
        for group_num, group_rules in groups.items():
            group_evals = []
            for rule in group_rules:
                eval_result = evaluate_condition(rule, client_data, entity_data)
                group_evals.append({
                    'condition': rule,
                    'evaluation': eval_result
                })

            # Within a group, use OR logic
            group_true = any(e['evaluation']['result'] == 'TRUE' for e in group_evals)
            group_unknown = any(e['evaluation']['result'] == 'UNKNOWN' for e in group_evals)

            if group_true:
                group_status = 'TRUE'
            elif group_unknown:
                group_status = 'UNKNOWN'
            else:
                group_status = 'FALSE'

            group_results.append({
                'group_num': group_num,
                'status': group_status,
                'evaluations': group_evals,
                'logic': group_rules[0].get('group_logic', 'OR')
            })

        # Between groups, apply the group logic (usually AND)
        # Simplified: if ANY group is TRUE and no groups are FALSE, requirement is triggered
        any_true = any(g['status'] == 'TRUE' for g in group_results)
        all_unknown = all(g['status'] == 'UNKNOWN' for g in group_results)

        if any_true:
            final_status = 'REQUIRED'
            confidence = 'HIGH'
        elif all_unknown:
            final_status = 'UNKNOWN'
            confidence = 'LOW - NEED DATA'
        else:
            final_status = 'NOT REQUIRED'
            confidence = 'HIGH'

        results.append({
            'rule_id': rule_id,
            'status': final_status,
            'confidence': confidence,
            'groups': group_results
        })

    # Overall: if ANY rule set triggers, requirement is met
    required = any(r['status'] == 'REQUIRED' for r in results)
    unknown = any(r['status'] == 'UNKNOWN' for r in results)

    if required:
        return {'required': True, 'confidence': 'HIGH', 'details': results}
    elif unknown:
        return {'required': 'LIKELY', 'confidence': 'LOW', 'details': results}
    else:
        return {'required': False, 'confidence': 'HIGH', 'details': results}


def assess_client(rules, client_data):
    """Generate compliance assessment for all client entities"""

    assessment = {}

    for entity in client_data['entities']:
        country = entity['country']

        # Filter rules for this country
        mf_rules = [r for r in rules['mf'] if r['country'] == country]
        lf_rules = [r for r in rules['lf'] if r['country'] == country]
        cbcr_rules = [r for r in rules['cbcr'] if r['country'] == country]

        # Evaluate requirements
        mf_result = evaluate_rule_set(mf_rules, client_data, entity)
        lf_result = evaluate_rule_set(lf_rules, client_data, entity)
        cbcr_result = evaluate_rule_set(cbcr_rules, client_data, entity)

        # Collect data gaps
        data_gaps = []
        for result in [mf_result, lf_result, cbcr_result]:
            for detail in result.get('details', []):
                for group in detail.get('groups', []):
                    for eval in group.get('evaluations', []):
                        if eval['evaluation']['result'] == 'UNKNOWN':
                            data_gaps.append({
                                'metric': eval['evaluation']['metric'],
                                'scope': eval['evaluation']['scope'],
                                'reason': eval['evaluation']['reason']
                            })

        assessment[entity['entity_name']] = {
            'country': country,
            'mf': mf_result,
            'lf': lf_result,
            'cbcr': cbcr_result,
            'data_gaps': data_gaps,
            'entity_data': entity
        }

    return assessment


def print_assessment(assessment, client_data):
    """Print assessment to console"""

    print("\n" + "="*80)
    print(f"TP COMPLIANCE ASSESSMENT - {client_data['client_info'].get('Client Name', 'Unknown Client')}")
    print(f"FYE: {client_data['client_info'].get('Fiscal Year End (FYE)', 'Unknown')}")
    print("="*80)

    for entity_name, result in assessment.items():
        print(f"\n{'─'*80}")
        print(f"ENTITY: {entity_name}")
        print(f"Country: {result['country']}")
        print(f"{'─'*80}")

        # Master File
        mf = result['mf']
        if mf['required'] == True:
            status = "✓ REQUIRED"
        elif mf['required'] == 'LIKELY':
            status = "⚠️  LIKELY REQUIRED - VERIFICATION NEEDED"
        else:
            status = "✗ NOT REQUIRED"

        print(f"\nMaster File: {status}")
        print(f"Confidence: {mf['confidence']}")

        for detail in mf.get('details', []):
            for group in detail.get('groups', []):
                for eval in group.get('evaluations', []):
                    result_icon = "✓" if eval['evaluation']['result'] == 'TRUE' else "?" if eval['evaluation']['result'] == 'UNKNOWN' else "✗"
                    print(f"  {result_icon} {eval['evaluation']['reason']}")

        # Local File
        lf = result['lf']
        if lf['required'] == True:
            status = "✓ REQUIRED"
        elif lf['required'] == 'LIKELY':
            status = "⚠️  LIKELY REQUIRED - VERIFICATION NEEDED"
        else:
            status = "✗ NOT REQUIRED"

        print(f"\nLocal File: {status}")
        print(f"Confidence: {lf['confidence']}")

        for detail in lf.get('details', []):
            for group in detail.get('groups', []):
                for eval in group.get('evaluations', []):
                    result_icon = "✓" if eval['evaluation']['result'] == 'TRUE' else "?" if eval['evaluation']['result'] == 'UNKNOWN' else "✗"
                    print(f"  {result_icon} {eval['evaluation']['reason']}")

        # CbCR
        cbcr = result['cbcr']
        if cbcr['required'] == True:
            status = "✓ REQUIRED"
        elif cbcr['required'] == 'LIKELY':
            status = "⚠️  LIKELY REQUIRED - VERIFICATION NEEDED"
        else:
            status = "✗ NOT REQUIRED"

        print(f"\nCbCR: {status}")
        print(f"Confidence: {cbcr['confidence']}")

        # Data Gaps
        if result['data_gaps']:
            print(f"\n📋 DATA NEEDED:")
            seen = set()
            for gap in result['data_gaps']:
                key = (gap['metric'], gap['scope'])
                if key not in seen:
                    print(f"  - {gap['metric']} ({gap['scope']})")
                    seen.add(key)

    print("\n" + "="*80)


def main():
    if len(sys.argv) < 3:
        print("Usage: python apply_rules.py Country_Rules_Library.xlsx Client_Data_Template.xlsx")
        sys.exit(1)

    rules_file = sys.argv[1]
    client_file = sys.argv[2]

    print(f"Reading rules from: {rules_file}")
    rules = read_rules(rules_file)

    print(f"Reading client data from: {client_file}")
    client_data = read_client_data(client_file)

    print("Applying rules...")
    assessment = assess_client(rules, client_data)

    print_assessment(assessment, client_data)


if __name__ == "__main__":
    main()
