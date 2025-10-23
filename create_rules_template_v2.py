#!/usr/bin/env python3
"""
Create IMPROVED Excel templates for rule-based TP compliance system
Version 2.0 - Combined rules and deadlines, better structure
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_country_rules_library_v2():
    """Create the improved Country Rules Library template"""

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Styling
    header_fill = PatternFill(start_color="0D2A5C", end_color="0D2A5C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    instruction_fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    example_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    section_fill = PatternFill(start_color="CFD8DC", end_color="CFD8DC", fill_type="solid")

    # ===== SHEET 1: Master File Requirements =====
    ws_mf = wb.create_sheet("MF Requirements")

    # Headers - Combined threshold rules and deadlines
    headers = [
        'Country',
        'Applicability',

        # THRESHOLD RULES section
        'Rule ID',
        'Condition Group',
        'Group Logic',
        'Metric Type',
        'Metric Scope',
        'Threshold Value',
        'Currency',
        'Operator',

        # DEADLINES section
        'Prep Date Rule',
        'Prep Date Details',
        'Submission Date Rule',
        'Submission Date Details',
        'Upon Request Days',

        # NOTES
        'Rule Notes',
        'Deadline Notes'
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws_mf.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Section headers
        if header in ['Rule ID', 'Condition Group']:
            cell.fill = section_fill
        elif header in ['Prep Date Rule', 'Submission Date Rule']:
            cell.fill = section_fill

    # Instructions row
    instructions = [
        'Germany, Spain, etc.',
        'Always / Conditional / Never Required / N/A',

        # THRESHOLD RULES
        'MF-DE-1, MF-ES-1',
        '1, 2, 3...',
        'OR / AND',
        'Revenue / Employees / RPTs / Always / etc.',
        'Group / Local Entity / Transaction',
        'Numeric threshold',
        'EUR / USD / etc.',
        '>= / > / = / < / <=',

        # DEADLINES
        'None / CIT Date / Fixed / FYE-Based',
        'Details if needed (e.g., CIT - 5 days)',
        'None / Upon Request / Fixed / FYE-Based',
        'Details (e.g., 2026-08-25 or FYE + 10 months)',
        'Days to submit if upon request (e.g., 30)',

        # NOTES
        'Threshold rule context',
        'Deadline context'
    ]

    for col, instr in enumerate(instructions, start=1):
        cell = ws_mf.cell(row=2, column=col)
        cell.value = instr
        cell.fill = instruction_fill
        cell.font = Font(italic=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Example: Germany MF
    examples = [
        # Country, Applicability, Rule ID, CondGrp, Logic, Metric, Scope, Threshold, Curr, Op, PrepRule, PrepDetails, SubRule, SubDetails, UponReqDays, RuleNotes, DeadlineNotes

        ['Germany', 'Conditional', 'MF-DE-1', '1', 'OR', 'Revenue', 'Group (Consolidated)', 750000000, 'EUR', '>=',
         'None', '', 'Upon Request', 'Within 30 days of audit notice', 30, 'Standard BEPS threshold', 'Automatic submission upon audit announcement'],

        ['Germany', 'Conditional', 'MF-DE-1', '2', 'OR', 'Revenue', 'Local Entity', 100000000, 'EUR', '>=',
         'None', '', 'Upon Request', 'Within 30 days of audit notice', 30, 'German entity sales threshold', ''],

        ['Spain', 'Conditional', 'MF-ES-1', '1', 'OR', 'Revenue', 'Group (Consolidated)', 45000000, 'EUR', '>',
         'CIT Date', 'Prepare by CIT filing (approx 25 Jul)', 'Upon Request', 'Within 10 days if audit', 10, 'Group consolidated turnover', 'Must be ready before CIT - 10 day audit window'],

        ['Malaysia', 'Never Required', '', '', '', '', '', '', '', '',
         '', '', '', '', '', 'MF content integrated into Local File (CTPD)', 'Group below threshold for standalone MF'],
    ]

    for row_idx, example in enumerate(examples, start=3):
        for col_idx, value in enumerate(example, start=1):
            cell = ws_mf.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.fill = example_fill
            if col_idx in [3, 11, 13]:  # Section starts
                cell.fill = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")

    # Set column widths
    ws_mf.column_dimensions['A'].width = 15  # Country
    ws_mf.column_dimensions['B'].width = 20  # Applicability
    ws_mf.column_dimensions['C'].width = 12  # Rule ID
    ws_mf.column_dimensions['D'].width = 12  # Condition Group
    ws_mf.column_dimensions['E'].width = 12  # Group Logic
    ws_mf.column_dimensions['F'].width = 20  # Metric Type
    ws_mf.column_dimensions['G'].width = 25  # Metric Scope
    ws_mf.column_dimensions['H'].width = 18  # Threshold
    ws_mf.column_dimensions['I'].width = 8   # Currency
    ws_mf.column_dimensions['J'].width = 8   # Operator
    ws_mf.column_dimensions['K'].width = 20  # Prep Date Rule
    ws_mf.column_dimensions['L'].width = 30  # Prep Date Details
    ws_mf.column_dimensions['M'].width = 20  # Submission Date Rule
    ws_mf.column_dimensions['N'].width = 30  # Submission Date Details
    ws_mf.column_dimensions['O'].width = 18  # Upon Request Days
    ws_mf.column_dimensions['P'].width = 35  # Rule Notes
    ws_mf.column_dimensions['Q'].width = 35  # Deadline Notes

    ws_mf.row_dimensions[1].height = 50
    ws_mf.row_dimensions[2].height = 80

    # ===== SHEET 2: Local File Requirements =====
    ws_lf = wb.create_sheet("LF Requirements")

    for col, header in enumerate(headers, start=1):
        cell = ws_lf.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if header in ['Rule ID', 'Condition Group', 'Prep Date Rule', 'Submission Date Rule']:
            cell.fill = section_fill

    for col, instr in enumerate(instructions, start=1):
        cell = ws_lf.cell(row=2, column=col)
        cell.value = instr
        cell.fill = instruction_fill
        cell.font = Font(italic=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Example: Germany & Spain LF
    examples_lf = [
        ['Germany', 'Conditional', 'LF-DE-1', '1', 'OR', 'RPTs', 'Transaction (Goods)', 6000000, 'EUR', '>',
         'None', '', 'Upon Request', 'Within 30 days of audit notice', 30, 'Goods-related transactions', 'LF submitted only upon request'],

        ['Germany', 'Conditional', 'LF-DE-1', '2', 'OR', 'RPTs', 'Transaction (Other)', 600000, 'EUR', '>',
         'None', '', 'Upon Request', 'Within 30 days of audit notice', 30, 'Other related party transactions', ''],

        ['Spain', 'Conditional', 'LF-ES-1', '1', 'OR', 'Revenue', 'Local Entity', 45000000, 'EUR', '>',
         'CIT Date', 'Prepare by CIT filing (approx 25 Jul)', 'Upon Request', 'Within 10 days if audit', 10, 'Net turnover threshold', 'Must be ready - 10 day audit response'],

        ['Spain', 'Conditional', 'LF-ES-1', '2', 'OR', 'RPTs', 'Transaction (All)', 250000, 'EUR', '>',
         'CIT Date', 'Prepare by CIT filing (approx 25 Jul)', 'Upon Request', 'Within 10 days if audit', 10, 'Related party transactions', ''],
    ]

    for row_idx, example in enumerate(examples_lf, start=3):
        for col_idx, value in enumerate(example, start=1):
            cell = ws_lf.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.fill = example_fill
            if col_idx in [3, 11, 13]:
                cell.fill = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")

    # Copy column widths
    for col in range(1, 18):
        ws_lf.column_dimensions[get_column_letter(col)].width = ws_mf.column_dimensions[get_column_letter(col)].width

    ws_lf.row_dimensions[1].height = 50
    ws_lf.row_dimensions[2].height = 80

    # ===== SHEET 3: CbCR Requirements =====
    ws_cbcr = wb.create_sheet("CbCR Requirements")

    for col, header in enumerate(headers, start=1):
        cell = ws_cbcr.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if header in ['Rule ID', 'Condition Group', 'Prep Date Rule', 'Submission Date Rule']:
            cell.fill = section_fill

    for col, instr in enumerate(instructions, start=1):
        cell = ws_cbcr.cell(row=2, column=col)
        cell.value = instr
        cell.fill = instruction_fill
        cell.font = Font(italic=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    examples_cbcr = [
        ['Germany', 'Conditional', 'CBCR-DE-1', '1', 'OR', 'Revenue', 'Group (Consolidated)', 750000000, 'EUR', '>=',
         'None', '', 'Fixed', '2025-12-31 (notification)', '', 'Standard OECD threshold', 'Notification via portal'],

        ['Spain', 'Conditional', 'CBCR-ES-1', '1', 'OR', 'Revenue', 'Group (Consolidated)', 750000000, 'EUR', '>=',
         'None', '', 'Fixed', '2025-12-31 (notification)', '', 'Standard OECD threshold', 'Annual notification'],
    ]

    for row_idx, example in enumerate(examples_cbcr, start=3):
        for col_idx, value in enumerate(example, start=1):
            cell = ws_cbcr.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.fill = example_fill
            if col_idx in [3, 11, 13]:
                cell.fill = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")

    for col in range(1, 18):
        ws_cbcr.column_dimensions[get_column_letter(col)].width = ws_mf.column_dimensions[get_column_letter(col)].width

    ws_cbcr.row_dimensions[1].height = 50
    ws_cbcr.row_dimensions[2].height = 80

    # ===== SHEET 4: TP Forms (Separate from Full Reports) =====
    ws_forms = wb.create_sheet("TP Forms")

    forms_headers = [
        'Country',
        'Form Type',
        'Form Name',
        'Form Triggers',
        'What It Contains',
        'Submission Date Rule',
        'Submission Date Details',
        'Electronic Signature Required?',
        'Timestamp Required?',
        'Notes'
    ]

    for col, header in enumerate(forms_headers, start=1):
        cell = ws_forms.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    forms_instructions = [
        'Country name',
        'TP Return / TP Disclosure / MF Summary / LF Summary / Notification',
        'Official form name/number',
        'Always / If MF Required / If LF Required / If Revenue > X',
        'Summary data / Full report / Checklist / etc.',
        'Fixed / FYE-Based / CIT-Based / With Tax Return',
        'Specific date or calculation (e.g., CIT + 30 days, 2026-08-25)',
        'Yes / No',
        'Yes / No / Electronic Timestamp',
        'Additional context'
    ]

    for col, instr in enumerate(forms_instructions, start=1):
        cell = ws_forms.cell(row=2, column=col)
        cell.value = instr
        cell.fill = instruction_fill
        cell.font = Font(italic=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    forms_examples = [
        ['Germany', 'TP Disclosure', 'Transaction Matrix', 'Always', 'Structured overview of RPTs', 'Upon Request', 'With MF upon audit (30 days)', 'No', 'Yes', 'NEW 2024 requirement - automatic with audit'],

        ['Spain', 'TP Return', 'Form 232', 'If MF or LF Required', 'Summary of TP info', 'Fixed', 'Approx 25 Aug 2026', 'Yes', 'No', 'Annual informative return'],

        ['Belgium', 'MF Summary', 'Form 275.MF', 'If MF Required', 'Summary form with MF data points', 'Fixed', '31 Dec 2026', 'Yes', 'No', 'Separate from full MF report'],

        ['Belgium', 'LF Summary', 'Form 275.LF', 'If LF Required', 'Summary form with LF data points', 'With Tax Return', 'Expected 31 Jul 2026', 'Yes', 'No', 'Separate from full LF report'],

        ['Belgium', 'Notification', 'Form 275.CBC.NOT', 'If CbCR Required', 'CbCR notification', 'Fixed', '31 Dec 2025', 'Yes', 'No', 'Annual CbCR notification'],

        ['Italy', 'TP Disclosure', 'RS 106', 'If MF or LF Required', 'Disclosure form in tax return', 'With Tax Return', 'Expected 31 Oct 2026', 'Yes', 'Yes - Electronic Timestamp', 'MF/LF must be timestamped before filing'],
    ]

    for row_idx, example in enumerate(forms_examples, start=3):
        for col_idx, value in enumerate(example, start=1):
            cell = ws_forms.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.fill = example_fill

    ws_forms.column_dimensions['A'].width = 15
    ws_forms.column_dimensions['B'].width = 20
    ws_forms.column_dimensions['C'].width = 25
    ws_forms.column_dimensions['D'].width = 25
    ws_forms.column_dimensions['E'].width = 30
    ws_forms.column_dimensions['F'].width = 25
    ws_forms.column_dimensions['G'].width = 35
    ws_forms.column_dimensions['H'].width = 20
    ws_forms.column_dimensions['I'].width = 20
    ws_forms.column_dimensions['J'].width = 40

    ws_forms.row_dimensions[1].height = 40
    ws_forms.row_dimensions[2].height = 80

    # ===== SHEET 5: Instructions =====
    ws_instructions = wb.create_sheet("Instructions", 0)

    instructions_text = """
TP COMPLIANCE RULES LIBRARY - VERSION 2.0
IMPROVED STRUCTURE - COMBINED RULES AND DEADLINES

═══════════════════════════════════════════════════════════════════════════════

WHAT'S NEW IN V2.0:

✓ Rules and deadlines on SAME sheet (no more separate Deadlines sheet)
✓ Applicability field: Always / Conditional / Never Required / N/A
✓ Three deadline types: Prep Date, Submission Date, Upon Request Days
✓ Separate sheet for TP Forms (distinct from full MF/LF reports)
✓ Timestamp and e-signature tracking

═══════════════════════════════════════════════════════════════════════════════

SHEET STRUCTURE:

SHEET 1: MF REQUIREMENTS
- When Master File documentation is required
- Threshold conditions (revenue, employees, etc.)
- Preparation and submission deadlines
- All on ONE sheet per country

SHEET 2: LF REQUIREMENTS
- When Local File documentation is required
- Same structure as MF

SHEET 3: CBCR REQUIREMENTS
- When Country-by-Country Reporting is required
- Usually just group revenue >= 750M EUR

SHEET 4: TP FORMS
- SEPARATE from full reports
- Forms like Belgium 275.MF (summary form) vs. full Master File report
- Forms like Spain Form 232 (informative return)
- Tracks e-signature and timestamp requirements

═══════════════════════════════════════════════════════════════════════════════

COLUMN REFERENCE - MF/LF/CBCR SHEETS:

COUNTRY: Country name (e.g., Germany, Spain)

APPLICABILITY:
  - Always: MF always required for this country
  - Conditional: MF required if thresholds met
  - Never Required: This country does not require MF (e.g., Malaysia MF)
  - N/A: Not applicable

THRESHOLD RULES SECTION:
  - Rule ID: Unique identifier (e.g., MF-DE-1)
  - Condition Group: Number (1, 2, 3...) for grouping related conditions
  - Group Logic: OR / AND between condition groups
  - Metric Type: Revenue / Employees / Balance Sheet / RPTs / Always
  - Metric Scope: Group / Local Entity / Transaction (Goods/Services/etc.)
  - Threshold Value: Numeric threshold
  - Currency: EUR, USD, etc.
  - Operator: >= / > / = / < / <=

DEADLINES SECTION:
  - Prep Date Rule: When documentation should be prepared
      * None: No specific prep deadline
      * CIT Date: Tie to Corporate Income Tax filing date
      * Fixed: Specific date every year
      * FYE-Based: Based on Fiscal Year End

  - Prep Date Details: Specifics
      * Examples: "CIT - 5 days", "FYE + 10 months", "By 31 Jul"

  - Submission Date Rule: When to submit if hard deadline exists
      * None: No hard deadline (only upon request)
      * Upon Request: Only submitted when authorities request
      * Fixed: Specific submission date
      * FYE-Based: Calculated from FYE

  - Submission Date Details: Specifics
      * Examples: "Within 30 days of audit notice", "2026-08-25", "FYE + 12 months"

  - Upon Request Days: Number of days to submit if requested
      * Example: 30 (submit within 30 days)
      * Leave blank if not applicable

NOTES:
  - Rule Notes: Context for threshold rules
  - Deadline Notes: Context for deadlines

═══════════════════════════════════════════════════════════════════════════════

DEADLINE PATTERNS EXPLAINED:

PATTERN 1: Upon Request Only (Germany MF/LF)
  - Prep Date Rule: None (no suggested prep date)
  - Submission Date Rule: Upon Request
  - Submission Date Details: Within 30 days of audit notice
  - Upon Request Days: 30

PATTERN 2: CIT-Based Prep, Upon Request Submission (Spain MF/LF)
  - Prep Date Rule: CIT Date
  - Prep Date Details: Prepare by CIT filing (approx 25 Jul)
  - Submission Date Rule: Upon Request
  - Submission Date Details: Within 10 days if audit
  - Upon Request Days: 10
  - Reason: Spain has 10-day audit response window - must be ready by CIT

PATTERN 3: Fixed Submission Deadline (Spain Form 232)
  - Prep Date Rule: CIT Date
  - Prep Date Details: By CIT filing date
  - Submission Date Rule: Fixed
  - Submission Date Details: Approx 25 Aug 2026
  - Upon Request Days: (blank)

PATTERN 4: FYE-Based Deadlines (India Form 3CEB)
  - Prep Date Rule: FYE-Based
  - Prep Date Details: FYE + 10 months
  - Submission Date Rule: Fixed
  - Submission Date Details: 30 Nov 2026
  - Upon Request Days: (blank)

PATTERN 5: With Tax Return (Belgium forms)
  - Prep Date Rule: CIT Date
  - Submission Date Rule: With Tax Return
  - Submission Date Details: Expected 31 Jul 2026

PATTERN 6: Electronic Timestamp Required (Italy)
  - Submission Date Rule: With Tax Return
  - Submission Date Details: Expected 31 Oct 2026
  - Notes: Must electronically timestamp MF/LF before filing RS 106

═══════════════════════════════════════════════════════════════════════════════

APPLICABILITY FIELD USAGE:

"Always"
  - Country always requires this (rare for MF/LF)
  - Example: Germany Transaction Matrix always required

"Conditional"
  - Most common - required if thresholds met
  - Define threshold conditions in same row

"Never Required"
  - This country does not have this requirement
  - Example: Malaysia Master File (MF content integrated into LF)
  - Leave threshold fields blank
  - Use Notes to explain why

"N/A"
  - Not applicable for this country
  - Country doesn't participate in this regime

═══════════════════════════════════════════════════════════════════════════════

TP FORMS VS. FULL REPORTS:

FULL REPORTS (MF/LF sheets):
  - Complete documentation packages
  - Master File report (50-100 pages)
  - Local File report (50-150 pages)
  - Comprehensive analysis

TP FORMS (TP Forms sheet):
  - Structured forms/schedules
  - Summary information
  - Data extracted from full reports
  - Examples:
    * Belgium Form 275.MF - summary form WITH MF data
    * Spain Form 232 - informative return
    * Germany Transaction Matrix - structured RPT overview
    * Italy RS 106 - disclosure in tax return

KEY DISTINCTION:
  Belgium example:
  - Full Master File Report: Complete MF documentation (100 pages)
  - Form 275.MF: Summary form filed with authorities (5 pages with key MF data)
  - Both required - form is separate from full report

═══════════════════════════════════════════════════════════════════════════════

MULTI-CONDITION LOGIC (UNCHANGED FROM V1.0):

Rules can have MULTIPLE CONDITIONS with AND/OR logic.

Example: Germany LF required if (Goods RPTs > 6M) OR (Other RPTs > 600K)

Row 1: LF-DE-1 | Group 1 | OR | RPTs | Goods | 6000000
Row 2: LF-DE-1 | Group 2 | OR | RPTs | Other | 600000

Same Rule ID, different Condition Groups, both use OR.

═══════════════════════════════════════════════════════════════════════════════

FILLING IN THE TEMPLATE:

FOR EACH COUNTRY:

Step 1: Determine Applicability
  - Is MF/LF ever required? → Conditional or Never Required
  - Is it always required? → Always
  - Not applicable? → N/A

Step 2: Define Threshold Rules (if Conditional)
  - What triggers the requirement?
  - Multiple conditions? Use same Rule ID, different groups

Step 3: Define Deadlines
  - When should it be prepared? (often CIT date)
  - Is there a hard submission deadline? (often no)
  - How many days if upon request? (often 30)

Step 4: Add Notes
  - Context for thresholds
  - Context for deadlines
  - Special requirements (e-signature, timestamp, etc.)

Step 5: Check TP Forms Sheet
  - Are there separate forms to file?
  - Add rows for each form
  - Note relationship to full reports

═══════════════════════════════════════════════════════════════════════════════

EXAMPLES IN THE TEMPLATE:

Germany:
  - MF: Conditional (revenue thresholds)
  - Prep: None
  - Submission: Upon Request (30 days from audit)

Spain:
  - MF: Conditional (revenue thresholds)
  - Prep: CIT Date (approx 25 Jul)
  - Submission: Upon Request (10 days - strict audit window)

Malaysia:
  - MF: Never Required (content in LF instead)

Belgium Forms:
  - Form 275.MF - separate summary form
  - Form 275.LF - separate summary form
  - Form 275.CBC.NOT - CbCR notification

Italy:
  - Electronic timestamp required for MF/LF before RS 106 filing

═══════════════════════════════════════════════════════════════════════════════

NEXT STEPS:

1. Review examples in MF/LF/CbCR sheets
2. Add more countries using same pattern
3. Fill in TP Forms sheet for countries with separate forms
4. Use with Client Data Template to run assessments

═══════════════════════════════════════════════════════════════════════════════

Version: 2.0 Prototype
Last Updated: 2025-10-23
Scope: Improved structure with combined rules and deadlines
"""

    ws_instructions['A1'] = instructions_text
    ws_instructions['A1'].alignment = Alignment(wrap_text=True, vertical='top')
    ws_instructions['A1'].font = Font(name='Consolas', size=10)
    ws_instructions.column_dimensions['A'].width = 120
    ws_instructions.row_dimensions[1].height = 2000

    wb.save('Country_Rules_Library_v2.xlsx')
    print("✓ Created: Country_Rules_Library_v2.xlsx")


if __name__ == "__main__":
    print("Creating IMPROVED Rule-Based TP Compliance Template (v2.0)...\n")
    create_country_rules_library_v2()
    print("\n✓ Template created successfully!")
    print("\nKEY IMPROVEMENTS:")
    print("  ✓ Rules and deadlines on SAME sheet")
    print("  ✓ Applicability field (Never Required option)")
    print("  ✓ Three deadline types: Prep / Submission / Upon Request")
    print("  ✓ Separate TP Forms sheet (forms vs. full reports)")
    print("  ✓ E-signature and timestamp tracking")
    print("\nNext: Open Country_Rules_Library_v2.xlsx and review the structure")
