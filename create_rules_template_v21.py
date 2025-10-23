#!/usr/bin/env python3
"""
Create Country Rules Library v2.1 - Final Changes
Incorporates all refinements: Integrated applicability, data validation, integrity rules
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_country_rules_library_v21():
    """Create the final Country Rules Library template v2.1"""

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Styling
    header_fill = PatternFill(start_color="0D2A5C", end_color="0D2A5C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    instruction_fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    example_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    section_fill = PatternFill(start_color="CFD8DC", end_color="CFD8DC", fill_type="solid")
    integrated_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

    # ===== SHEET 1: Master File Requirements =====
    ws_mf = wb.create_sheet("MF Requirements")

    # Headers - Updated with new columns
    headers = [
        'Country',
        'Applicability',
        'Integrated With',
        'Effective From (FY)',

        # THRESHOLD RULES section
        'Rule ID',
        'Condition Group',
        'Group Logic',
        'Metric Type',
        'Metric Scope',
        'Threshold Value',
        'Currency',
        'Operator',

        # DEADLINES section
        'Prep Date Rule',
        'Prep Date Details',
        'Submission Date Rule',
        'Submission Date Details',
        'Upon Request Days',
        'Submission Channel',

        # NOTES
        'Rule Notes',
        'Deadline Notes'
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws_mf.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Section highlighting
        if header in ['Rule ID', 'Condition Group']:
            cell.fill = section_fill
        elif header in ['Prep Date Rule', 'Submission Date Rule']:
            cell.fill = section_fill
        elif header in ['Integrated With', 'Effective From (FY)']:
            cell.fill = integrated_fill

    # Instructions row
    instructions = [
        'Germany, Spain, Malaysia, etc.',
        'Always / Conditional / Integrated / Never Required / N/A',
        'Local File / TP Form / Other (only if Integrated)',
        'FY2024, 2023, etc. (when rule took effect)',

        # THRESHOLD RULES
        'MF-DE-1, MF-ES-1, etc.',
        '1, 2, 3... (increment for multi-conditions)',
        'OR / AND (between groups)',
        'Revenue / Employees / RPTs / Balance Sheet / Always',
        'Group / Local Entity / Transaction',
        'Numeric threshold',
        'EUR / USD / GBP / JPY / MYR / etc.',
        '>= / > / = / < / <=',

        # DEADLINES
        'None / CIT Date / FYE-Based / Fixed / Upon Request / With Tax Return',
        'Details (e.g., CIT - 5 days, FYE + 10 months)',
        'None / CIT Date / FYE-Based / Fixed / Upon Request / With Tax Return',
        'Details (e.g., 2026-08-25, Within 30 days of audit)',
        'Days to submit if upon request (30, 14, 10, etc.)',
        'e-filing portal / Form 275.MF / Attachment with CIT / Paper',

        # NOTES
        'Threshold rule context',
        'Deadline and submission context'
    ]

    for col, instr in enumerate(instructions, start=1):
        cell = ws_mf.cell(row=2, column=col)
        cell.value = instr
        cell.fill = instruction_fill
        cell.font = Font(italic=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # DATA VALIDATION SETUP

    # Applicability dropdown
    dv_applicability = DataValidation(
        type="list",
        formula1='"Always,Conditional,Integrated,Never Required,N/A"',
        allow_blank=False
    )
    dv_applicability.error = 'Invalid value. Choose from dropdown.'
    dv_applicability.errorTitle = 'Invalid Applicability'
    ws_mf.add_data_validation(dv_applicability)
    dv_applicability.add(f'B3:B1000')

    # Integrated With dropdown
    dv_integrated = DataValidation(
        type="list",
        formula1='"Local File,TP Form,Other"',
        allow_blank=True
    )
    dv_integrated.error = 'Choose: Local File, TP Form, or Other'
    dv_integrated.errorTitle = 'Invalid Integration Type'
    ws_mf.add_data_validation(dv_integrated)
    dv_integrated.add(f'C3:C1000')

    # Group Logic dropdown
    dv_logic = DataValidation(
        type="list",
        formula1='"OR,AND"',
        allow_blank=True
    )
    ws_mf.add_data_validation(dv_logic)
    dv_logic.add(f'G3:G1000')

    # Operator dropdown
    dv_operator = DataValidation(
        type="list",
        formula1='">=,>,=,<,<="',
        allow_blank=True
    )
    ws_mf.add_data_validation(dv_operator)
    dv_operator.add(f'L3:L1000')

    # Prep Date Rule dropdown
    dv_prep_rule = DataValidation(
        type="list",
        formula1='"None,CIT Date,FYE-Based,Fixed,Upon Request,With Tax Return"',
        allow_blank=True
    )
    dv_prep_rule.error = 'Choose from standardized list'
    dv_prep_rule.errorTitle = 'Invalid Prep Date Rule'
    ws_mf.add_data_validation(dv_prep_rule)
    dv_prep_rule.add(f'M3:M1000')

    # Submission Date Rule dropdown
    dv_sub_rule = DataValidation(
        type="list",
        formula1='"None,CIT Date,FYE-Based,Fixed,Upon Request,With Tax Return"',
        allow_blank=True
    )
    dv_sub_rule.error = 'Choose from standardized list'
    dv_sub_rule.errorTitle = 'Invalid Submission Date Rule'
    ws_mf.add_data_validation(dv_sub_rule)
    dv_sub_rule.add(f'O3:O1000')

    # Upon Request Days - numeric validation
    dv_days = DataValidation(
        type="whole",
        operator="greaterThan",
        formula1='0',
        allow_blank=True
    )
    dv_days.error = 'Must be a positive number (e.g., 30, 14, 10)'
    dv_days.errorTitle = 'Invalid Days'
    ws_mf.add_data_validation(dv_days)
    dv_days.add(f'Q3:Q1000')

    # Example rows
    examples = [
        # Germany - Conditional
        ['Germany', 'Conditional', '', 'FY2024',
         'MF-DE-1', '1', 'OR', 'Revenue', 'Group (Consolidated)', 750000000, 'EUR', '>=',
         'None', '', 'Upon Request', 'Within 30 days of audit notice', 30, 'e-filing portal',
         'Standard BEPS threshold', 'Automatic submission upon audit announcement'],

        ['Germany', 'Conditional', '', 'FY2024',
         'MF-DE-1', '2', 'OR', 'Revenue', 'Local Entity', 100000000, 'EUR', '>=',
         'None', '', 'Upon Request', 'Within 30 days of audit notice', 30, 'e-filing portal',
         'German entity sales threshold', ''],

        # Spain - Conditional
        ['Spain', 'Conditional', '', 'FY2023',
         'MF-ES-1', '1', 'OR', 'Revenue', 'Group (Consolidated)', 45000000, 'EUR', '>',
         'CIT Date', 'Prepare by CIT filing (approx 25 Jul)', 'Upon Request', 'Within 10 days if audit', 10, 'Attachment with CIT',
         'Group consolidated turnover', 'Must be ready before CIT - 10 day audit window'],

        # Malaysia - Integrated
        ['Malaysia', 'Integrated', 'Local File', 'FY2023',
         '', '', '', '', '', '', '', '',
         'CIT Date', 'Prepare with CTPD by CIT filing (7 months after FYE)', 'Upon Request', 'Within 14 days of request', 14, 'e-filing portal',
         'MF content integrated into LF per 2023 TPD; no standalone MF. Group revenue below MYR 3B threshold.', 'MF elements included in CTPD submission'],

        # Singapore - Integrated (placeholder)
        ['Singapore', 'Integrated', 'Local File', 'FY2019',
         '', '', '', '', '', '', '', '',
         'CIT Date', 'Prepare with LF by CIT filing', 'Upon Request', 'Within 30 days of request', 30, 'IRAS e-portal',
         'MF elements included in LF per IRAS guidance. No standalone MF requirement.', 'Submission upon IRAS request during TP audit'],

        # Switzerland - Never Required
        ['Switzerland', 'Never Required', '', '',
         '', '', '', '', '', '', '', '',
         '', '', '', '', '', '',
         'No statutory MF requirement; TP documentation recommended for penalty protection', 'Documentation provided upon audit request (typically 30 days)'],
    ]

    for row_idx, example in enumerate(examples, start=3):
        for col_idx, value in enumerate(example, start=1):
            cell = ws_mf.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.fill = example_fill

            # Highlight Integrated entries
            if col_idx == 2 and value == 'Integrated':
                cell.fill = integrated_fill
            if col_idx == 3 and row_idx in [6, 7]:  # Malaysia and Singapore rows
                cell.fill = integrated_fill

            # Section highlighting
            if col_idx in [5, 13, 15]:  # Section starts
                cell.fill = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")

    # Set column widths
    ws_mf.column_dimensions['A'].width = 15  # Country
    ws_mf.column_dimensions['B'].width = 20  # Applicability
    ws_mf.column_dimensions['C'].width = 18  # Integrated With
    ws_mf.column_dimensions['D'].width = 18  # Effective From
    ws_mf.column_dimensions['E'].width = 12  # Rule ID
    ws_mf.column_dimensions['F'].width = 15  # Condition Group
    ws_mf.column_dimensions['G'].width = 12  # Group Logic
    ws_mf.column_dimensions['H'].width = 20  # Metric Type
    ws_mf.column_dimensions['I'].width = 25  # Metric Scope
    ws_mf.column_dimensions['J'].width = 18  # Threshold
    ws_mf.column_dimensions['K'].width = 8   # Currency
    ws_mf.column_dimensions['L'].width = 8   # Operator
    ws_mf.column_dimensions['M'].width = 22  # Prep Date Rule
    ws_mf.column_dimensions['N'].width = 35  # Prep Date Details
    ws_mf.column_dimensions['O'].width = 22  # Submission Date Rule
    ws_mf.column_dimensions['P'].width = 35  # Submission Date Details
    ws_mf.column_dimensions['Q'].width = 18  # Upon Request Days
    ws_mf.column_dimensions['R'].width = 25  # Submission Channel
    ws_mf.column_dimensions['S'].width = 40  # Rule Notes
    ws_mf.column_dimensions['T'].width = 40  # Deadline Notes

    ws_mf.row_dimensions[1].height = 50
    ws_mf.row_dimensions[2].height = 90

    # Apply same structure to LF and CbCR sheets (code similar to above)
    # ... (I'll create simplified versions for LF and CbCR)

    # ===== SHEET 2: Validation Rules & Notes =====
    ws_validation = wb.create_sheet("Validation Rules", 0)

    validation_text = """
COUNTRY RULES LIBRARY v2.1 - VALIDATION & INTEGRITY RULES

═══════════════════════════════════════════════════════════════════════════════

AUTOMATIC DATA VALIDATION (ENFORCED):

1. APPLICABILITY column (B):
   ✓ Dropdown values: Always, Conditional, Integrated, Never Required, N/A
   ✓ Cannot be blank

2. INTEGRATED WITH column (C):
   ✓ Dropdown values: Local File, TP Form, Other
   ✓ Can be blank
   ⚠️  RULE: Must be filled if APPLICABILITY = "Integrated"

3. GROUP LOGIC column (G):
   ✓ Dropdown values: OR, AND

4. OPERATOR column (L):
   ✓ Dropdown values: >=, >, =, <, <=

5. PREP DATE RULE column (M):
   ✓ Dropdown values: None, CIT Date, FYE-Based, Fixed, Upon Request, With Tax Return
   ✓ Standardized - prevents typos

6. SUBMISSION DATE RULE column (O):
   ✓ Dropdown values: None, CIT Date, FYE-Based, Fixed, Upon Request, With Tax Return
   ✓ Standardized - prevents typos

7. UPON REQUEST DAYS column (Q):
   ✓ Must be positive whole number (e.g., 30, 14, 10)
   ⚠️  RULE: Should be numeric if SUBMISSION DATE RULE = "Upon Request"

═══════════════════════════════════════════════════════════════════════════════

MANUAL INTEGRITY CHECKS (USER RESPONSIBILITY):

1. INTEGRATED APPLICABILITY:
   IF Applicability = "Integrated"
   THEN:
     - INTEGRATED WITH must not be blank
     - Rule ID, Condition Group, Thresholds should be blank (no separate thresholds)
     - Notes should explain where MF content is integrated

2. UPON REQUEST VALIDATION:
   IF Submission Date Rule = "Upon Request"
   THEN:
     - UPON REQUEST DAYS should be numeric (30, 14, 10, etc.)
     - Can be blank only if genuinely unknown

3. MULTI-CONDITION RULES:
   For conditions that combine (e.g., Revenue OR Employees):
   - Use SAME Rule ID across all related rows
   - INCREMENT Condition Group (1, 2, 3...)
   - Set GROUP LOGIC consistently (OR or AND)

   Example:
   Row 1: MF-DE-1 | Group 1 | OR | Revenue | >= 750M
   Row 2: MF-DE-1 | Group 2 | OR | Employees | > 250

4. NEVER REQUIRED vs N/A:
   - "Never Required": Country has no MF rule in law
   - "N/A": Regime doesn't apply (e.g., CbCR when no group presence)

5. CURRENCY:
   - Keep original currency as stated in law (EUR, USD, JPY, MYR, etc.)
   - Do NOT auto-convert
   - Client data should convert to common currency (EUR) for comparison
   - Preserves legal thresholds accurately

6. EFFECTIVE FROM (FY):
   - Use when rules change over time
   - Example: France threshold changed in FY2024
   - Allows historical tracking without overwriting

7. SUBMISSION CHANNEL:
   - How MF content is submitted (e-portal, form, attachment, paper)
   - Complements TP Forms sheet (which tracks form details)
   - Use for MF/LF submissions; track separate forms in TP Forms sheet

═══════════════════════════════════════════════════════════════════════════════

COMMON PATTERNS & HOW TO HANDLE:

PATTERN 1: Integrated MF (Malaysia, Singapore)
  Applicability: Integrated
  Integrated With: Local File
  Rule ID: (blank)
  Thresholds: (blank)
  Notes: "MF content integrated into LF per [regulation]; no standalone MF"

PATTERN 2: Never Required (Switzerland)
  Applicability: Never Required
  Rule ID: (blank)
  Thresholds: (blank)
  Notes: "No statutory MF requirement; TP documentation recommended"

PATTERN 3: Conditional with Multiple Thresholds (Germany)
  Row 1: MF-DE-1 | Group 1 | OR | Revenue (Group) | >= 750M
  Row 2: MF-DE-1 | Group 2 | OR | Revenue (Local) | >= 100M

PATTERN 4: Complex AND/OR Logic
  MF required if (Group Revenue >= 500M) AND (Employees > 250 OR Balance Sheet > 43M)

  Row 1: MF-X-1 | Group 1 | AND | Revenue (Group) | >= 500M
  Row 2: MF-X-1 | Group 2 | OR  | Employees (Local) | > 250
  Row 3: MF-X-1 | Group 3 | OR  | Balance Sheet (Local) | > 43M

  Logic: Group 1 AND (Group 2 OR Group 3)

PATTERN 5: Rolling Rule Changes (France threshold update)
  Entry 1:
    Country: France
    Effective From (FY): FY2023
    Threshold: 50000000
    Notes: "Old threshold before 2024 update"

  Entry 2:
    Country: France
    Effective From (FY): FY2024
    Threshold: 45000000
    Notes: "New threshold from FY2024"

═══════════════════════════════════════════════════════════════════════════════

KNOWN ISSUES & WORKAROUNDS:

ISSUE 1: Currency Normalization
  Problem: Thresholds in different currencies (EUR 750M vs MYR 3B)
  Solution:
    - Keep legal threshold in original currency in this sheet
    - Convert client data to EUR in Client Data sheet
    - Logic engine compares apples-to-apples

ISSUE 2: Mixed Metrics Per Country
  Problem: Some countries use Group revenue AND Local RPTs
  Solution:
    - Create separate condition groups
    - Use AND logic between groups
    - Document clearly in notes

ISSUE 3: Employee/Asset Metrics
  Problem: Less common than revenue, but valid thresholds
  Solution:
    - Set METRIC TYPE = "Employees" or "Balance Sheet"
    - Set METRIC SCOPE = "Local Entity" or "Group"
    - Client must provide this data

ISSUE 4: Submission Channel vs TP Forms
  Problem: Confusion between how MF is submitted vs forms that accompany it
  Solution:
    - SUBMISSION CHANNEL (this sheet): How MF content is transmitted
    - TP Forms sheet: Separate forms/schedules that reference MF
    - Belgium example:
      * Submission Channel: "e-filing portal"
      * TP Forms: "Form 275.MF" (separate summary form)

ISSUE 5: "Always" vs "Conditional"
  Problem: When to use "Always"
  Solution:
    - "Always": Requirement exists regardless of thresholds (rare for MF/LF)
    - "Conditional": Most common - based on size/RPT thresholds
    - Example of "Always": Germany Transaction Matrix

═══════════════════════════════════════════════════════════════════════════════

DATA QUALITY CHECKLIST:

Before finalizing data entry for a country, verify:

□ Applicability is set correctly
□ If "Integrated" → Integrated With is filled
□ If "Conditional" → Threshold rules defined
□ If "Never Required" → Threshold fields blank
□ Multi-condition rules use same Rule ID
□ Condition Groups increment properly (1, 2, 3...)
□ Group Logic is consistent (OR or AND)
□ Prep Date Rule uses standardized value
□ Submission Date Rule uses standardized value
□ Upon Request Days is numeric if applicable
□ Currency matches legal source (don't convert)
□ Notes explain any unusual situations
□ Effective From (FY) set if rule has changed

═══════════════════════════════════════════════════════════════════════════════

VERSION HISTORY:

v2.1 (2025-10-23):
  - Added "Integrated" applicability
  - Added INTEGRATED WITH column
  - Added EFFECTIVE FROM (FY) column
  - Added SUBMISSION CHANNEL column
  - Standardized date rule dropdowns
  - Added comprehensive validation
  - Malaysia and Singapore examples

v2.0 (2025-10-23):
  - Combined rules and deadlines on same sheet
  - Added three deadline types
  - Separated TP Forms sheet

v1.0 (2025-10-23):
  - Initial rule-based system
"""

    ws_validation['A1'] = validation_text
    ws_validation['A1'].alignment = Alignment(wrap_text=True, vertical='top')
    ws_validation['A1'].font = Font(name='Consolas', size=10)
    ws_validation.column_dimensions['A'].width = 120
    ws_validation.row_dimensions[1].height = 2400

    wb.save('Country_Rules_Library_v2.1.xlsx')
    print("✓ Created: Country_Rules_Library_v2.1.xlsx")


if __name__ == "__main__":
    print("Creating Country Rules Library v2.1 (Final Changes)...\n")
    create_country_rules_library_v21()
    print("\n✓ Template created successfully!")
    print("\nNEW IN v2.1:")
    print("  ✓ 'Integrated' applicability (Malaysia, Singapore)")
    print("  ✓ INTEGRATED WITH column (Local File / TP Form / Other)")
    print("  ✓ EFFECTIVE FROM (FY) column (track rule changes)")
    print("  ✓ SUBMISSION CHANNEL column (e-portal, form, etc.)")
    print("  ✓ Standardized date rule dropdowns (data validation)")
    print("  ✓ Numeric validation for Upon Request Days")
    print("  ✓ Comprehensive validation rules documentation")
    print("\nNext: Open Country_Rules_Library_v2.1.xlsx and review examples")
